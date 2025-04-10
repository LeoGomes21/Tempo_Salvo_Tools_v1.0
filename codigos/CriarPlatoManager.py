from qgis.core import QgsProject, QgsRasterLayer, QgsMapSettings, QgsMapRendererCustomPainterJob, Qgis, QgsMessageLog, QgsWkbTypes, QgsVectorLayer, QgsPoint, QgsField, QgsFeature, QgsGeometry, QgsPointXY, QgsRaster, QgsMarkerSymbol, QgsLineSymbol, QgsArrowSymbolLayer, QgsVectorLayerSimpleLabeling, QgsPalLayerSettings, QgsTextFormat, QgsProperty, QgsCoordinateTransform, QgsLineString, QgsTextBackgroundSettings, QgsSpatialIndex
from PyQt5.QtWidgets import QDialog, QCheckBox, QComboBox, QPushButton, QGraphicsView, QGraphicsScene, QGraphicsPixmapItem, QGraphicsTextItem, QScrollArea, QLabel, QLineEdit, QDoubleSpinBox, QVBoxLayout, QHBoxLayout, QLabel, QListWidgetItem, QListWidget, QFileDialog, QStyledItemDelegate, QStyle
from PyQt5.QtGui import QImage, QPainter, QPixmap, QColor, QFont, QPen, QBrush
from qgis.PyQt.QtCore import Qt, QRectF, QPointF, QSize, QVariant, QRect, QPoint, QEvent
from matplotlib.offsetbox import OffsetImage, AnnotationBbox
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
from PyQt5.QtGui import QDoubleValidator
from matplotlib.patches import Circle
import matplotlib.patches as patches
from openpyxl import load_workbook
from qgis.gui import QgsMapCanvas
from PyQt5.QtCore import QLocale
import matplotlib.pyplot as plt
from qgis.utils import iface
from qgis.PyQt import uic
import pandas as pd
import numpy as np
import tempfile
import math
import os

FORM_CLASS, _ = uic.loadUiType(os.path.join(
    os.path.dirname(__file__), 'platoMDT.ui'))

class PlatoManager(QDialog, FORM_CLASS):
    def __init__(self, parent=None):
        """Constructor."""
        super(PlatoManager, self).__init__(parent)
        # Configura a interface do usuário a partir do Designer.
        self.setupUi(self)

        self.manual_edit_flags = {}  # Dicionário para sinalizar IDs editados manualmente

        # Altera o título da janela
        self.setWindowTitle("Gerar Platô sobre o MDT")

        # Inicializa o sinalizador para controlar a execução do gráfico
        self.is_generating_graph = False

        # Cria uma cena gráfica para o QGraphicsView
        self.scene = QGraphicsScene()
        self.graphicsViewRaster.setScene(self.scene)

        self.scenePoligono = QGraphicsScene()
        self.graphicsViewPoligono.setScene(self.scenePoligono)

        # Inicializa o ComboBox de Raster
        self.init_combo_box_raster()
        self.init_combo_box_poligono()

        self.scroll_widgets = []  # Armazena os widgets da scrollArea

        # Cria uma instância do GraficoManager e passa a instância atual de PlatoManager
        self.grafico_manager = GraficoManager(self)

        # Inicializa o caminho do logotipo como None
        self.logo_path = None

        # Configurações iniciais
        self.check_logo_status()  # Verifica e ajusta o estado do botão de logo

        # Conecta os sinais aos slots
        self.connect_signals()

        # Configura a janela para permitir minimizar, maximizar e fechar
        self.setWindowFlags(self.windowFlags() | Qt.WindowMinimizeButtonHint | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)

    def connect_signals(self):

        # Conecta os sinais existentes
        self.comboBoxRaster.currentIndexChanged.connect(self.display_raster)
        self.comboBoxPoligono.currentIndexChanged.connect(self.display_polygon)  # Conexão para atualizar o graphicsViewPoligono
        QgsProject.instance().layersAdded.connect(self.handle_layers_added)

        # Conectar o novo método que lida com a remoção de camadas
        QgsProject.instance().layersRemoved.connect(self.handle_layers_removed)

        for layer in QgsProject.instance().mapLayers().values():
            layer.nameChanged.connect(self.update_combo_boxes)

        # Conectar sinais relacionados a outras verificações que podem afetar o botão
        for layer in QgsProject.instance().mapLayers().values():
            layer.nameChanged.connect(self.update_pushButtonVertices_state)

        # Conecta o botão pushButtonCalcular para atualizar as camadas
        self.pushButtonCalcular.clicked.connect(self.calculate_and_create_segment_layers)

        # Atualiza a lista de segmentos ao abrir o projeto ou ao carregar camadas
        QgsProject.instance().layersAdded.connect(self.update_list_widget_segmentos)
        QgsProject.instance().layerWillBeRemoved.connect(self.update_list_widget_segmentos)

        # Atualiza a lista de Estaqueamento ao abrir o projeto ou ao carregar camadas
        QgsProject.instance().layersAdded.connect(self.update_list_widget_estaqueamentos)
        QgsProject.instance().layerWillBeRemoved.connect(self.update_list_widget_estaqueamentos)

        # Atualiza a lista de Taludes ao abrir o projeto ou ao carregar camadas
        QgsProject.instance().layersAdded.connect(self.update_list_widget_talude)
        QgsProject.instance().layerWillBeRemoved.connect(self.update_list_widget_talude)

        # Conecta o botão para calcular e gerar os pontos intermediários
        self.pushButtonEstaqueamento.clicked.connect(self.calculate_intermediate_points)

        # Conecta o botão para calcular e gerar os Taludes
        self.pushButtonTalude.clicked.connect(self.calculate_talude)

        # Conecta a seleção de itens no listWidgetSegmentos
        self.listWidgetSegmentos.itemSelectionChanged.connect(self.check_estaqueamento_button_state)

        # Conecta o valor do doubleSpinBoxEquidistante para ativar/desativar o botão
        self.doubleSpinBoxEquidistante.valueChanged.connect(self.check_estaqueamento_button_state)

        # Conectar o sinal do doubleSpinBoxEquidistante para atualizar o toolTip
        self.doubleSpinBoxEquidistante.valueChanged.connect(self.calculate_and_update_tooltip)

        # Conectar o sinal de seleção do listWidgetSegmentos para recalcular o toolTip
        self.listWidgetSegmentos.currentItemChanged.connect(self.calculate_and_update_tooltip)

        # Conectar o sinal de seleção do listWidgetSegmentos para recalcular o toolTip e ajustar o estado do doubleSpinBox
        self.listWidgetSegmentos.currentItemChanged.connect(self.update_doubleSpinBoxEquidistante_state)

        # Conectar o sinal de seleção do listWidgetEstacas para ajustar o estado do pushButtonTalude
        self.listWidgetEstacas.currentItemChanged.connect(self.update_pushButtonTalude_state)

        # Conecta as mudanças no listWidgetEstacas
        self.listWidgetEstacas.itemSelectionChanged.connect(self.update_radio_buttons_state)

        # Conecta as mudanças no listWidgetEstacas para ativar ou desativar o pushButtonGrafico
        self.listWidgetEstacas.itemSelectionChanged.connect(self.update_pushButtonGrafico_state)

        # Atualiza o estado do pushButtonVertices ao mudar a seleção no comboBoxPoligono
        self.comboBoxPoligono.currentIndexChanged.connect(self.update_pushButtonVertices_state)
        self.comboBoxRaster.currentIndexChanged.connect(self.update_pushButtonVertices_state)

        # Atualiza o estado do pushButtonJuntar na alteração do listWidgetEstacas e listWidgetTalude
        self.listWidgetEstacas.itemSelectionChanged.connect(self.update_pushButtonJuntar_state)
        self.listWidgetTalude.itemSelectionChanged.connect(self.update_pushButtonJuntar_state)

        # Conectar eventos relacionados a mudanças no projeto
        QgsProject.instance().layersAdded.connect(self.update_pushButtonVertices_state)
        QgsProject.instance().layersRemoved.connect(self.update_pushButtonVertices_state)

        # Desconectar quaisquer conexões existentes para evitar múltiplas conexões
        try:
            self.pushButtonGrafico.clicked.disconnect()
            self.pushButtonJuntar.clicked.disconnect()
            self.pushButtonLogo.clicked.disconnect()
            self.pushButtonVertices.clicked.disconnect()
        except TypeError:
            pass  # Nenhuma conexão existente

        # Conecta o botão para criar a camada de pontos a partir dos vértices
        self.pushButtonVertices.clicked.connect(self.create_point_layer_from_polygon)

        # Conecta o botão para calcular e gerar o Gráfico
        self.pushButtonGrafico.clicked.connect(self.on_pushButtonGrafico_clicked)

        # Conecta o botão para juntar as camadas
        self.pushButtonJuntar.clicked.connect(self.on_pushButtonJuntar_clicked)

        # Conectar o botão pushButtonCancelar para fechar o diálogo
        self.pushButtonCancelar.clicked.connect(self.close)

        # Conectar o botão pushButtonLogo para abrir a pasta
        self.pushButtonLogo.clicked.connect(self.on_pushButtonLogo_clicked)

        # Inicializar o botão "Calcular" como desativado
        self.pushButtonCalcular.setEnabled(False)

        # Conectar o botão "Vertices" ao método que ativará o botão "Calcular"
        self.pushButtonVertices.clicked.connect(self.on_pushButtonVertices_clicked)

        # Certifique-se de que a scrollAreaWidgetContents tenha um layout
        if not self.scrollAreaWidgetContents.layout():
            layout = QVBoxLayout()  # ou outro tipo de layout conforme necessário
            self.scrollAreaWidgetContents.setLayout(layout)

        # Instanciar o delegado e definir nos listWidgets
        self.delegateSegmentos = ListDeleteButtonDelegate(self.listWidgetSegmentos)
        self.listWidgetSegmentos.setItemDelegate(self.delegateSegmentos)

        self.delegateEstacas = ListDeleteButtonDelegate(self.listWidgetEstacas)
        self.listWidgetEstacas.setItemDelegate(self.delegateEstacas)

        self.delegateTalude = ListDeleteButtonDelegate(self.listWidgetTalude)
        self.listWidgetTalude.setItemDelegate(self.delegateTalude)

        # Conecte o evento de duplo clique ao listWidgetEstacas
        self.listWidgetEstacas.itemDoubleClicked.connect(self.open_excel_with_attributes)

    def init_combo_box_raster(self):
        # Armazenar o índice atual selecionado
        current_raster_id = self.comboBoxRaster.currentData()

        layers = QgsProject.instance().mapLayers().values()
        raster_layers = [layer for layer in layers if layer.type() == layer.RasterLayer]
        self.comboBoxRaster.clear()
        for raster_layer in raster_layers:
            self.comboBoxRaster.addItem(raster_layer.name(), raster_layer.id())

        if current_raster_id:
            index = self.comboBoxRaster.findData(current_raster_id)
            if index != -1:
                self.comboBoxRaster.setCurrentIndex(index)  # Restaura a seleção anterior
            else:
                self.comboBoxRaster.setCurrentIndex(0)  # Seleciona o primeiro item se o anterior não existir
        else:
            if raster_layers:
                self.comboBoxRaster.setCurrentIndex(0)

        self.display_raster()

    def init_combo_box_poligono(self):
        # Inicializa o ComboBox de Polígono
        layers = QgsProject.instance().mapLayers().values()
        polygon_layers = [layer for layer in layers if layer.type() == layer.VectorLayer and layer.geometryType() == QgsWkbTypes.PolygonGeometry]
        self.comboBoxPoligono.clear()
        for polygon_layer in polygon_layers:
            self.comboBoxPoligono.addItem(polygon_layer.name(), polygon_layer.id())

            # Conecta os sinais featuresAdded e featuresDeleted para camadas de polígono existentes
            polygon_layer.featureAdded.connect(self.display_polygon)
            polygon_layer.featureDeleted.connect(self.display_polygon)

        if polygon_layers:
            self.comboBoxPoligono.setCurrentIndex(0)
            self.display_polygon()

    def display_raster(self):
        # Limpa a cena antes de adicionar um novo item
        self.scene.clear()

        # Obtém o ID da camada raster selecionada
        selected_raster_id = self.comboBoxRaster.currentData()

        # Busca a camada raster pelo ID
        selected_layer = QgsProject.instance().mapLayer(selected_raster_id)

        if isinstance(selected_layer, QgsRasterLayer):
            # Configurações do mapa
            map_settings = QgsMapSettings()
            map_settings.setLayers([selected_layer])  # Definimos a camada a ser renderizada
            map_settings.setBackgroundColor(QColor(255, 255, 255))
            
            # Define o tamanho da imagem a ser renderizada
            width = self.graphicsViewRaster.viewport().width()
            height = self.graphicsViewRaster.viewport().height()
            map_settings.setOutputSize(QSize(width, height))
            
            # Define a extensão do mapa (extensão do raster)
            map_settings.setExtent(selected_layer.extent())

            # Cria a imagem para renderizar
            image = QImage(width, height, QImage.Format_ARGB32)
            image.fill(Qt.transparent)

            # Configura o pintor e a tarefa de renderização
            painter = QPainter(image)
            render_job = QgsMapRendererCustomPainterJob(map_settings, painter)

            # Executa a renderização
            render_job.start()
            render_job.waitForFinished()
            painter.end()

            # Cria um pixmap a partir da imagem renderizada
            pixmap = QPixmap.fromImage(image)
            pixmap_item = QGraphicsPixmapItem(pixmap)

            # Adiciona o item à cena
            self.scene.addItem(pixmap_item)

            # Ajusta a cena ao QGraphicsView, garantindo que o modo de ajuste preserve a proporção
            self.graphicsViewRaster.setSceneRect(pixmap_item.boundingRect())
            self.graphicsViewRaster.fitInView(self.scene.sceneRect(), Qt.KeepAspectRatio)

    def display_polygon(self):
        """Atualiza a exibição do polígono no QGraphicsView com base na seleção de feições."""
        self.scenePoligono.clear()
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        
        if selected_layer and isinstance(selected_layer, QgsVectorLayer):
            features = [f for f in selected_layer.getFeatures() if f.geometry().type() == QgsWkbTypes.PolygonGeometry]

            # Verifica se há feições na camada
            if len(features) == 0:
                if self.isVisible():  # Verifica se o diálogo está aberto
                    self.mostrar_mensagem("A camada de polígono não contém feições.", "Erro")
                return

            # Verifica se há feições selecionadas
            selected_features = selected_layer.selectedFeatures()
            if len(selected_features) > 0:
                feature = selected_features[0]  # Exibe a feição selecionada
            else:
                # Se não houver feições selecionadas, usa a primeira feição da camada
                if len(features) > 0:
                    feature = features[0]
                else:
                    if self.isVisible():  # Verifica se o diálogo está aberto
                        self.mostrar_mensagem("Nenhuma feição válida encontrada na camada.", "Erro")
                    return

            # Verifica se a feição tem geometria válida
            geom = feature.geometry()
            if geom is None or geom.isEmpty():
                if self.isVisible():  # Verifica se o diálogo está aberto
                    self.mostrar_mensagem("A feição selecionada não contém geometria válida.", "Erro")
                return

            # Renderiza a geometria da feição no QGraphicsView
            map_settings = QgsMapSettings()
            map_settings.setLayers([selected_layer])
            map_settings.setBackgroundColor(QColor(255, 255, 255))
            width = self.graphicsViewPoligono.viewport().width()
            height = self.graphicsViewPoligono.viewport().height()
            map_settings.setOutputSize(QSize(width, height))
            
            # Verifica se o boundingBox da geometria é válido
            bounding_box = geom.boundingBox()
            if not bounding_box.isNull():
                map_settings.setExtent(bounding_box)
            else:
                if self.isVisible():  # Verifica se o diálogo está aberto
                    self.mostrar_mensagem("A geometria não tem um bounding box válido.", "Erro")
                return

            # Criação da imagem para renderizar
            image = QImage(width, height, QImage.Format_ARGB32)
            image.fill(Qt.transparent)
            painter = QPainter(image)
            render_job = QgsMapRendererCustomPainterJob(map_settings, painter)
            render_job.start()
            render_job.waitForFinished()
            painter.end()

            # Exibir no QGraphicsView
            pixmap = QPixmap.fromImage(image)
            pixmap_item = QGraphicsPixmapItem(pixmap)
            self.scenePoligono.addItem(pixmap_item)
            self.graphicsViewPoligono.setSceneRect(pixmap_item.boundingRect())
            self.graphicsViewPoligono.fitInView(self.scenePoligono.sceneRect(), Qt.KeepAspectRatio)
        else:
            # self.mostrar_mensagem ("Nenhuma camada de polígono válida foi encontrada.", "Erro")
            pass

    def get_z_value(self, x, y):
        selected_raster_id = self.comboBoxRaster.currentData()
        raster_layer = QgsProject.instance().mapLayer(selected_raster_id)

        if not raster_layer or not isinstance(raster_layer, QgsRasterLayer):
            self.mostrar_mensagem(
                "Camada raster inválida ou não selecionada.",
                "Erro"
            )
            return None

        identify_result = raster_layer.dataProvider().identify(
            QgsPointXY(x, y), 
            QgsRaster.IdentifyFormatValue
        )

        if identify_result.isValid():
            results = identify_result.results()
            if results:
                band_key = list(results.keys())[0]
                z_value = results.get(band_key)  # <-- Pode ser None
                if z_value is None:
                    self.mostrar_mensagem(
                        f"Não há valor de elevação (Z) para o pixel em ({x}, {y}). "
                        "Verifique se o polígono está dentro do raster.",
                        "Erro"
                    )
                    return None
                # Se `z_value` for válido, convertemos para float
                return round(float(z_value), 3)

        # Caso identify_result não seja válido ou esteja vazio
        self.mostrar_mensagem(
            f"Não foi possível identificar o valor Z em ({x}, {y}). "
            "Verifique se o polígono está dentro do raster.",
            "Erro"
        )
        return None

    def _log_message(self, message, level=Qgis.Info):
        QgsMessageLog.logMessage(message, 'PLATÔ', level=level)

    def mostrar_mensagem(self, texto, tipo, duracao=1):
        """
        Exibe uma mensagem na barra de mensagens da interface do QGIS.

        Parâmetros:
        texto: str - O texto da mensagem a ser exibida.
        tipo: str - O tipo da mensagem, que pode ser "Erro" ou "Sucesso".
        duracao: int - A duração da mensagem em segundos (padrão é 3 segundos).

        Funções:
        - Obtém a barra de mensagens da interface do QGIS.
        - Exibe uma mensagem de erro ou sucesso com o nível apropriado baseado no tipo fornecido.
        """
        
        bar = iface.messageBar()  # Acessa a barra de mensagens da interface do QGIS

        # Exibe a mensagem com o nível apropriado baseado no tipo
        if tipo == "Erro":
            # Mostra uma mensagem de erro na barra de mensagens com um ícone crítico e a duração especificada
            bar.pushMessage("Erro", texto, level=Qgis.Critical, duration=duracao)
        elif tipo == "Sucesso":
            # Cria o item da mensagem
            msg = bar.createMessage("Sucesso", texto)
            
            # Adiciona a mensagem à barra com o nível informativo e a duração especificada
            bar.pushWidget(msg, level=Qgis.Info, duration=duracao)

    def reset_scroll_area(self):
        """Reseta e limpa a área de rolagem (scrollArea) e a lista de widgets (scroll_widgets)."""
        # Limpa a lista de widgets armazenada
        self.scroll_widgets.clear()

        # Limpa o layout do scrollArea
        layout = self.scrollAreaWidgetContents.layout()
        
        if layout is not None:
            while layout.count() > 0:
                item = layout.takeAt(0)
                if item.widget() is not None:
                    item.widget().deleteLater()
                elif item.layout() is not None:
                    self.clear_layout(item.layout())  # Limpa layouts aninhados

    def clear_layout(self, layout):
        """Limpa todos os widgets e layouts de um layout específico."""
        while layout.count() > 0:
            item = layout.takeAt(0)
            if item.widget() is not None:
                item.widget().deleteLater()
            elif item.layout() is not None:
                self.clear_layout(item.layout())

    def showEvent(self, event):
        super(PlatoManager, self).showEvent(event)

        # Limpa a lista de widgets no scrollArea
        self.reset_scroll_area()

        # Ajusta a visualização quando o diálogo é mostrado
        self.display_raster()
        self.display_polygon()

        # Atualiza o listWidgetSegmentos ao abrir a janela
        self.update_list_widget_segmentos()
        self.update_list_widget_estaqueamentos()
        self.update_list_widget_talude()

        # Reseta o estado do botão pushButtonEstaqueamento (desativado)
        self.pushButtonEstaqueamento.setEnabled(False)

        # Reseta o valor do doubleSpinBoxEquidistante
        self.doubleSpinBoxEquidistante.setValue(0)

        # Define o radioButtonEsquerda como ativo por padrão
        self.radioButtonEsquerda.setChecked(True)

        # Verifica e aplica as configurações de ativação do pushButtonVertices
        self.update_pushButtonVertices_state()

        # Atualiza o estado do doubleSpinBoxEquidistante baseado no listWidgetSegmentos
        self.update_doubleSpinBoxEquidistante_state()

        # Atualiza o estado do pushButtonTalude baseado no listWidgetEstacas
        self.update_pushButtonTalude_state()
        
        # Atualiza o estado do radioButtonEsquerda e radioButtonDireita baseado no listWidgetEstacas
        self.update_radio_buttons_state()
        
        # Atualiza o estado do pushButtonGrafico baseado no listWidgetEstacas
        self.update_pushButtonGrafico_state()

        # Atualiza o estado do pushButtonJuntar baseado no listWidgetEstacas ou listWidgetTalude
        self.update_pushButtonJuntar_state()

        # Atualiza o estado do pushButtonCalcular baseado no click do botão pushButtonVertices
        self.on_pushButtonVertices_clicked()

        self.connect_selection_signal() # Chama após carregar ou selecionar a camada

    def closeEvent(self, event):
        parent = self.parent()
        if parent:
            parent.plato_mdt_dlg = None
        super(PlatoManager, self).closeEvent(event)

    def update_combo_box(self):
        # Armazena a seleção atual para restaurar após a atualização
        current_index = self.comboBoxRaster.currentIndex()
        current_layer_id = self.comboBoxRaster.itemData(current_index)

        # Atualiza o combo box quando camadas são adicionadas ou removidas
        self.init_combo_box_raster()

        # Tenta restaurar a seleção anterior
        if current_layer_id:
            index = self.comboBoxRaster.findData(current_layer_id)
            if index != -1:
                self.comboBoxRaster.setCurrentIndex(index)
            else:
                # Se a camada não existe mais, seleciona a primeira disponível
                if self.comboBoxRaster.count() > 0:
                    self.comboBoxRaster.setCurrentIndex(0)
                    self.display_raster()

    def update_combo_boxes(self):
        # Atualiza ambos os combo boxes
        self.init_combo_box_raster()
        self.init_combo_box_poligono()

    def handle_layers_added(self, layers):
        # Salva qual polígono está selecionado no momento
        current_polygon_id = self.comboBoxPoligono.currentData()
        
        # Atualiza comboBoxes normalmente
        self.update_combo_boxes()
        
        # Tenta restaurar a seleção do polígono anterior
        if current_polygon_id is not None:
            index = self.comboBoxPoligono.findData(current_polygon_id)
            if index != -1:
                self.comboBoxPoligono.setCurrentIndex(index)

        # Agora verifica se foi adicionado algum layer de polígono
        polygon_layers = [layer for layer in layers 
                          if isinstance(layer, QgsVectorLayer) 
                          and layer.geometryType() == QgsWkbTypes.PolygonGeometry]

        if polygon_layers:
            self.display_polygon()

            # Conecta sinais para atualizar visualização sempre que 
            # feições forem adicionadas/removidas
            for polygon_layer in polygon_layers:
                polygon_layer.featureAdded.connect(self.display_polygon)
                polygon_layer.featureDeleted.connect(self.display_polygon)

    def handle_layers_removed(self, removed_layers_ids):
        """Função para tratar a remoção de camadas."""

        # Atualizar os ComboBoxes de Raster e Polígono
        self.update_combo_boxes()

        # Verificar e atualizar os grupos e widgets associados
        self.check_layers_in_group_and_update(removed_layers_ids)

    def apply_arrow_symbology(self, layer):
        """Aplica simbologia de seta a uma camada de linhas."""
        # Cria o símbolo de linha
        line_symbol = QgsLineSymbol.createSimple({'color': 'blue', 'width': '0.5'})

        # Cria uma camada de símbolo de linha que inclui uma seta
        arrow_layer = QgsArrowSymbolLayer()
        arrow_layer.setHeadLength(5)  # Define o tamanho da cabeça da seta
        arrow_layer.setHeadThickness(2)  # Define a espessura da cabeça da seta
        arrow_layer.setArrowWidth(1)  # Define a largura da seta

        # Adiciona a camada de símbolo de seta ao símbolo de linha
        line_symbol.appendSymbolLayer(arrow_layer)

        # Define a simbologia na camada
        layer.renderer().setSymbol(line_symbol)
        layer.triggerRepaint()

    def activate_labels_for_points(self, point_layer):
        """Ativa os rótulos de 'NovoZ' na camada de pontos com formatação simples, em negrito e cor preta, com deslocamento."""
        
        layer_settings = QgsPalLayerSettings()
        text_format = QgsTextFormat()

        # Configurações básicas do rótulo
        layer_settings.isExpression = True  # Definimos como expressão para combinar 'ID' e 'NovoZ'
        layer_settings.fieldName = "'ID: ' || \"ID\" || '\nNovoZ: ' || \"NovoZ\""  # Expressão para exibir ID e NovoZ com quebra de linha
        layer_settings.placement = QgsPalLayerSettings.OverPoint  # Coloca o rótulo sobre o ponto

        # Configurações do texto do rótulo
        font = QFont("Arial", 10)
        font.setBold(True)  # Negrito simples
        text_format.setFont(font)
        text_format.setSize(10)  # Tamanho do texto

        # Definir a cor do texto (preto)
        text_format.setColor(QColor(0, 0, 0))  # Cor preta para o rótulo

        # Aplicar o formato de texto ao rótulo
        layer_settings.setFormat(text_format)

        # Configuração de deslocamento (offset) do rótulo
        layer_settings.xOffset = 6  # Deslocamento horizontal
        layer_settings.yOffset = 3  # Deslocamento vertical

        # Definir a rotulagem na camada
        labeling = QgsVectorLayerSimpleLabeling(layer_settings)
        point_layer.setLabeling(labeling)
        point_layer.setLabelsEnabled(True)
        point_layer.triggerRepaint()

    def activate_labels_on_layer(self, layer, field_name):
        """Ativa os rótulos em uma camada de linhas com símbolos de seta no rótulo e formatação condicional."""
        
        layer_settings = QgsPalLayerSettings()
        text_format = QgsTextFormat()

        # Configurações básicas do rótulo
        layer_settings.isExpression = True  # Definimos como expressão para incluir setas
        layer_settings.fieldName = f"CASE WHEN \"{field_name}\" < 0 THEN '⬇' || round(\"{field_name}\" * 100, 3) || '%' " \
                                   f"WHEN \"{field_name}\" > 0 THEN '⬆' || round(\"{field_name}\" * 100, 3) || '%' " \
                                   f"ELSE '→' || round(\"{field_name}\" * 100, 3) || '%' END"  # Inclui setas com base no valor de inclinação

        layer_settings.placement = QgsPalLayerSettings.Line  # Coloca o rótulo ao longo da linha

        # Configurações do texto do rótulo
        font = QFont("Arial", 14)
        font.setBold(True)
        font.setItalic(True)
        text_format.setFont(font)
        text_format.setSize(14)

        # Configuração do fundo branco para o rótulo
        background = QgsTextBackgroundSettings()
        background.setEnabled(True)
        background.setFillColor(QColor(255, 255, 0))  # Cor de fundo amarelo
        text_format.setBackground(background)

        # Adiciona formatação condicional (vermelho para negativo, azul para positivo, preto para zero)
        color_expression = """CASE
                                WHEN "{field}" < 0 THEN '255,0,0'  -- Vermelho para inclinações negativas
                                WHEN "{field}" > 0 THEN '0,0,255'  -- Azul para inclinações positivas
                                ELSE '0,0,0'  -- Preto para zero
                              END""".format(field=field_name)
        
        # Definir as propriedades dos rótulos (cor, tamanho)
        properties = layer_settings.dataDefinedProperties()
        properties.setProperty(QgsPalLayerSettings.Color, QgsProperty.fromExpression(color_expression))
        properties.setProperty(QgsPalLayerSettings.Size, QgsProperty.fromValue(10))  # Tamanho da fonte

        # Aplicar o formato de texto ao rótulo
        layer_settings.setFormat(text_format)
        layer_settings.setDataDefinedProperties(properties)

        # Definir a rotulagem na camada
        labeling = QgsVectorLayerSimpleLabeling(layer_settings)
        layer.setLabeling(labeling)
        layer.setLabelsEnabled(True)
        layer.triggerRepaint()

    def create_point_layer_from_polygon(self):
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        
        if selected_layer and isinstance(selected_layer, QgsVectorLayer):
            features = [f for f in selected_layer.getFeatures() if f.geometry().type() == QgsWkbTypes.PolygonGeometry]

            # Verifica se há apenas uma feição ou uma feição selecionada
            if len(features) == 1 or selected_layer.selectedFeatureCount() == 1:
                if selected_layer.selectedFeatureCount() == 1:
                    feature = selected_layer.selectedFeatures()[0]
                else:
                    feature = features[0]

                # Nome da camada de pontos com base na camada de polígono
                point_layer_name = f"{selected_layer.name()}_Vértices"

                # Verifica se já existe uma camada com o mesmo nome
                existing_layers = QgsProject.instance().mapLayersByName(point_layer_name)
                if existing_layers:
                    # Tenta remover a camada existente
                    try:
                        QgsProject.instance().removeMapLayer(existing_layers[0].id())
                    except Exception as e:
                        self.mostrar_mensagem(f"Não foi possível sobrescrever a camada existente: {str(e)}", "Erro")
                        return

                # Cria a camada de pontos
                point_layer = QgsVectorLayer('Point?crs=' + selected_layer.crs().authid(), point_layer_name, 'memory')
                pr = point_layer.dataProvider()

                # Adiciona campos "ID", "X", "Y", "Z", "NovoZ"
                pr.addAttributes([QgsField("ID", QVariant.Int),
                                  QgsField("X", QVariant.Double),
                                  QgsField("Y", QVariant.Double),
                                  QgsField("Z", QVariant.Double),
                                  QgsField("NovoZ", QVariant.Double)])
                point_layer.updateFields()

                # Obtém o raster selecionado para calcular o valor de Z
                selected_raster_id = self.comboBoxRaster.currentData()
                raster_layer = QgsProject.instance().mapLayer(selected_raster_id)

                if not raster_layer or not isinstance(raster_layer, QgsRasterLayer):
                    self.mostrar_mensagem("Camada raster inválida ou não selecionada.", "Erro")
                    return

                vertices = feature.geometry().asPolygon()[0]

                # Remove o último vértice se for igual ao primeiro para evitar duplicação
                if vertices[0] == vertices[-1]:
                    vertices.pop()

                points = []
                for i, vertex in enumerate(vertices):
                    x = round(vertex.x(), 3)
                    y = round(vertex.y(), 3)

                    # Identifica o valor no raster
                    identify_result = raster_layer.dataProvider().identify(QgsPointXY(x, y), QgsRaster.IdentifyFormatValue)

                    # Verifica se foi válido e se há resultados
                    if identify_result.isValid():
                        results = identify_result.results()
                        if results:
                            band_key = list(results.keys())[0]
                            z_val = results[band_key]
                            
                            # Aqui está o ponto principal: verificar se o z_val é None
                            if z_val is None:
                                self.mostrar_mensagem(
                                    f"Nenhum valor de elevação (Z) encontrado em ({x}, {y}). Verifique se seu polígono está dentro do raster.",
                                    "Erro"
                                )
                                # Decida se quer pular esse ponto (continue) ou encerrar o método (return).
                                continue  
                            
                            z = round(float(z_val), 3)
                        else:
                            self.mostrar_mensagem(
                                f"Não há resultados de elevação para ({x}, {y}). Verifique se seu polígono está dentro do raster.",
                                "Erro")
                            continue
                    else:
                        self.mostrar_mensagem(
                            f"Não foi possível identificar valor de elevação para ({x}, {y}). Polígono pode estar fora do raster.",
                            "Erro")
                        continue

                    # Cria o ponto
                    point = QgsFeature(point_layer.fields())
                    point.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(x, y)))
                    point.setAttribute("ID", i + 1)
                    point.setAttribute("X", x)
                    point.setAttribute("Y", y)
                    point.setAttribute("Z", z)
                    point.setAttribute("NovoZ", z)
                    points.append(point)

                    # Adiciona o ponto à camada de pontos
                    pr.addFeature(point)

                # Adiciona a camada de pontos ao projeto
                QgsProject.instance().addMapLayer(point_layer)
                self.mostrar_mensagem("Camada de pontos criada com sucesso.", "Sucesso")

                # Ativar os rótulos de 'NovoZ' na camada de pontos
                self.activate_labels_for_points(point_layer)

                # Criar a camada de linhas, incluindo a linha que fecha o polígono
                self.create_line_layer_from_points(points, selected_layer.crs())

                # Popula a scrollArea com os dados das camadas
                self.populate_scroll_area()

            else:
                self.mostrar_mensagem("Selecione apenas 1 feição", "Erro")

    def create_line_layer_from_points(self, points, crs):
        # Obter o nome da camada de polígono selecionada
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        if not selected_layer:
            # self.mostrar_mensagem("'.", "Erro")
            return
        
        # Definir o nome da camada de linhas baseado na camada de polígono
        line_layer_name = f"{selected_layer.name()}_Segmentos"

        # Verifica se já existe uma camada com o mesmo nome
        existing_layers = QgsProject.instance().mapLayersByName(line_layer_name)
        if existing_layers:
            # Tenta remover a camada existente
            try:
                QgsProject.instance().removeMapLayer(existing_layers[0].id())
            except Exception as e:
                self.mostrar_mensagem(f"Não foi possível sobrescrever a camada existente: {str(e)}", "Erro")
                return

        # Cria a camada de linhas
        line_layer = QgsVectorLayer(f'LineString?crs={crs.authid()}', line_layer_name, 'memory')
        pr = line_layer.dataProvider()

        # Adiciona campos "ID", "Comprimento", "Inclinação"
        pr.addAttributes([QgsField("ID", QVariant.Int),
                          QgsField("Comprimento", QVariant.Double),
                          QgsField("Inclinação", QVariant.Double)])
        line_layer.updateFields()

        num_points = len(points)

        for i in range(num_points):
            point1 = points[i]
            point2 = points[(i + 1) % num_points]  # Conecta o último ponto ao primeiro

            # Cria a linha entre dois pontos consecutivos
            line = QgsFeature(line_layer.fields())
            line_geom = QgsGeometry.fromPolylineXY([point1.geometry().asPoint(), point2.geometry().asPoint()])
            comprimento = round(line_geom.length(), 3)

            # Calcula a inclinação
            z1 = point1['NovoZ']
            z2 = point2['NovoZ']
            inclinacao = round((float(z2) - float(z1)) / comprimento if comprimento != 0 else 0, 5)

            # Atribui os valores à feição
            line.setGeometry(line_geom)
            line.setAttribute("ID", i + 1)
            line.setAttribute("Comprimento", comprimento)
            line.setAttribute("Inclinação", inclinacao)

            # Adiciona a linha à camada de linhas
            pr.addFeature(line)

        # Atribui a simbologia de seta à camada de linhas
        self.apply_arrow_symbology(line_layer)

        # Ativa os rótulos na camada de linhas (usando o campo "Inclinação" como exemplo)
        self.activate_labels_on_layer(line_layer, "Inclinação")

        # Adiciona a camada de linhas ao projeto
        QgsProject.instance().addMapLayer(line_layer)
        self.mostrar_mensagem("Camada de linhas criada com sucesso.", "Sucesso")

    def populate_scroll_area(self):
        # Reset the scroll area before setting it up again
        self.reset_scroll_area()

        # Obter o nome da camada de polígono selecionada
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)

        if not selected_layer:
            # self.mostrar_mensagem("Camada de polígono não encontrada.", "Erro")
            return
        
        # Definir o nome da camada de pontos baseada na camada de polígono
        point_layer_name = f"{selected_layer.name()}_Vértices"

        # Definir o nome da camada de linhas baseada na camada de polígono
        line_layer_name = f"{selected_layer.name()}_Segmentos"

        # Obter as camadas de pontos e linhas pelo nome dinâmico
        point_layer = QgsProject.instance().mapLayersByName(point_layer_name)
        line_layer = QgsProject.instance().mapLayersByName(line_layer_name)

        if not point_layer or not line_layer:
            self.mostrar_mensagem(f"Camadas '{point_layer_name}' ou 'Linhas' não encontradas.", "Erro")
            return

        point_layer = point_layer[0]  # Obtém a primeira camada correspondente (espera-se que só haja uma)
        line_layer = line_layer[0]

        points = []
        lines = []

        # Extrair dados da camada de pontos
        for feature in point_layer.getFeatures():
            point_data = {
                'ID': feature['ID'],
                'Z': feature['Z'],
                'NovoZ': feature['NovoZ'],
            }
            points.append(point_data)

        # Extrair dados da camada de linhas
        for feature in line_layer.getFeatures():
            line_data = {
                'ID': feature['ID'],
                'Inclinação': feature['Inclinação'],
                'Comprimento': feature['Comprimento'],
            }
            lines.append(line_data)

        # Popula a scroll area com os dados obtidos
        self.setup_scroll_area(points, lines)

    def setup_scroll_area(self, points, lines):
        self.points = points  # Armazena os pontos para uso posterior
        self.lines = lines  # Armazena as linhas para uso posterior

        # Certifica-se de que o layout da scrollArea está configurado
        layout = self.scrollAreaWidgetContents.layout()
        if layout is None:
            layout = QVBoxLayout(self.scrollAreaWidgetContents)
            self.scrollAreaWidgetContents.setLayout(layout)
        else:
            # Limpa a scroll area
            while layout.count() > 0:
                item = layout.takeAt(0)
                if item.widget():
                    item.widget().deleteLater()

        self.scroll_widgets = []  # Armazena referências aos widgets para fácil acesso

        # Adiciona cabeçalhos
        header_layout = QHBoxLayout()
        headers = ["ID's", "Z", "NovoZ", "Inclinação", "Desnível"]
        
        for header in headers:
            label = QLabel(header)
            label.setAlignment(Qt.AlignLeft)  # Centraliza o texto do cabeçalho
            label.setStyleSheet("font-weight: bold;")  # Negrito para o cabeçalho
            header_layout.addWidget(label)
        
        layout.addLayout(header_layout)

        # Cria as linhas de inputs para cada ponto
        for i, point in enumerate(points):
            hbox = QHBoxLayout()

            # QLabel para o ID
            label_id = QLabel(f"ID: {i + 1}")
            label_id.setAlignment(Qt.AlignLeft)  # Centraliza o texto do ID
            hbox.addWidget(label_id)

            # QLineEdit para o Z (não editável)
            line_edit_z = QLineEdit()
            line_edit_z.setText(str(point['Z']))
            line_edit_z.setReadOnly(True)
            line_edit_z.setAlignment(Qt.AlignLeft)  # Centraliza o texto
            hbox.addWidget(line_edit_z)

            # QLineEdit para o NovoZ (editável)
            line_edit_novoz = QLineEdit()
            line_edit_novoz.setText(str(point['NovoZ']))
            line_edit_novoz.setAlignment(Qt.AlignLeft)  # Centraliza o texto à esquerda

            # Configura o QDoubleValidator para aceitar apenas números
            validator = QDoubleValidator(-99999.999, 99999.999, 3, line_edit_novoz)
            validator.setNotation(QDoubleValidator.StandardNotation)
            validator.setLocale(QLocale(QLocale.English))  # Define a localidade para usar '.' como separador decimal
            line_edit_novoz.setValidator(validator)

            line_edit_novoz.editingFinished.connect(lambda idx=i: self.update_novoz(idx))
            hbox.addWidget(line_edit_novoz)

            # QLineEdit para a Inclinação (não editável)
            line_edit_inclinacao = QLineEdit()
            if i < len(lines):
                line_edit_inclinacao.setText(str(lines[i]['Inclinação']))
            line_edit_inclinacao.setReadOnly(True)
            line_edit_inclinacao.setAlignment(Qt.AlignLeft)  # Centraliza o texto À esquerda
            hbox.addWidget(line_edit_inclinacao)

            # QDoubleSpinBox para o Desnível (editável)
            spin_box_desnivel = QDoubleSpinBox()
            spin_box_desnivel.setDecimals(3)
            spin_box_desnivel.setMinimum(-9999.99)
            spin_box_desnivel.setMaximum(9999.99)
            spin_box_desnivel.setSingleStep(0.5)
            desnivel = point['NovoZ'] - point['Z']

            spin_box_desnivel.setValue(desnivel)  # Valor inicial baseado no NovoZ
            spin_box_desnivel.setAlignment(Qt.AlignLeft) # Centraliza o texto
            spin_box_desnivel.valueChanged.connect(lambda val, idx=i: self.update_desnivel(val, idx))
            hbox.addWidget(spin_box_desnivel)

            # Armazena os widgets para fácil acesso posterior
            self.scroll_widgets.append({
                'line_edit_novoz': line_edit_novoz,
                'line_edit_inclinacao': line_edit_inclinacao,
                'spin_box_desnivel': spin_box_desnivel
            })

            # Agora que o dicionário já está na lista, conecte o sinal
            spin_box_desnivel.valueChanged.connect(lambda val, idx=i: self.update_desnivel(val, idx))

            # E o mesmo para o line_edit_novoz, se necessário
            line_edit_novoz.editingFinished.connect(lambda idx=i: self.update_novoz(idx))

            layout.addLayout(hbox)

        layout.addStretch()  # Adiciona um esticador para garantir que o conteúdo fique no topo

    def recalculate_inclination_and_desnivel(self):
        num_points = len(self.points)

        for i in range(num_points):
            # Recalcula a inclinação para o segmento anterior (inclui a conexão circular para o primeiro ponto)
            if i > 0:
                z1 = self.points[i - 1]['NovoZ']
                z2 = self.points[i]['NovoZ']
                comprimento = self.lines[i - 1]['Comprimento']
                inclinacao = round((z2 - z1) / comprimento if comprimento != 0 else 0, 3)
                
                # Atualiza o QLineEdit de inclinação para o segmento anterior
                self.scroll_widgets[i - 1]['line_edit_inclinacao'].setText(str(inclinacao))
            else:
                # Para o primeiro ponto, recalcula a inclinação com o último ponto
                z1 = self.points[-1]['NovoZ']
                z2 = self.points[i]['NovoZ']
                comprimento = self.lines[-1]['Comprimento']
                inclinacao = round((z2 - z1) / comprimento if comprimento != 0 else 0, 3)
                
                # Atualiza o QLineEdit de inclinação para o último segmento
                self.scroll_widgets[-1]['line_edit_inclinacao'].setText(str(inclinacao))

            # Recalcula a inclinação para o segmento posterior (inclui a conexão circular para o último ponto)
            if i < num_points - 1:
                z1 = self.points[i]['NovoZ']
                z2 = self.points[i + 1]['NovoZ']
                comprimento = self.lines[i]['Comprimento']
                inclinacao = round((z2 - z1) / comprimento if comprimento != 0 else 0, 3)

                # Atualiza o QLineEdit de inclinação para o segmento posterior
                self.scroll_widgets[i]['line_edit_inclinacao'].setText(str(inclinacao))
            else:
                # Para o último ponto, recalcula a inclinação com o primeiro ponto
                z1 = self.points[i]['NovoZ']
                z2 = self.points[0]['NovoZ']
                comprimento = self.lines[i]['Comprimento']
                inclinacao = round((z2 - z1) / comprimento if comprimento != 0 else 0, 3)

                # Atualiza o QLineEdit de inclinação para o primeiro segmento
                self.scroll_widgets[i]['line_edit_inclinacao'].setText(str(inclinacao))

            # Recalcula o desnível
            z_orig = self.points[i]['Z']
            z_novo = self.points[i]['NovoZ']
            desnivel = round(z_novo - z_orig, 3)

            # Atualiza o valor do QDoubleSpinBox correspondente
            self.scroll_widgets[i]['spin_box_desnivel'].setValue(desnivel)

    def update_desnivel(self, val, idx):
        try:
            desnivel = float(val)
        except ValueError:
            return

        # Calcula o novo valor de NovoZ com base no desnível e arredonda para três casas decimais
        novo_z = round(self.points[idx]['Z'] + desnivel, 3)

        # Atualiza o NovoZ no ponto específico
        self.points[idx]['NovoZ'] = novo_z

        # Atualiza o QLineEdit correspondente ao NovoZ para refletir a alteração com 3 casas decimais
        self.scroll_widgets[idx]['line_edit_novoz'].setText(f"{novo_z:.3f}")

        # Recalcula a inclinação e o desnível para os segmentos adjacentes
        self.recalculate_inclination_and_desnivel()

    def update_novoz(self, idx):
        line_edit_novoz = self.scroll_widgets[idx]['line_edit_novoz']
        
        try:
            # Obtenha o valor digitado no QLineEdit
            novo_z = float(line_edit_novoz.text())
        except ValueError:
            return

        # Arredonda o valor para três casas decimais
        novo_z = round(novo_z, 3)

        # Atualiza o NovoZ no ponto específico
        self.points[idx]['NovoZ'] = novo_z

        # Atualiza o QLineEdit correspondente ao NovoZ para refletir a alteração com 3 casas decimais
        line_edit_novoz.setText(f"{novo_z:.3f}")

        # Recalcula a inclinação e o desnível para os segmentos adjacentes
        self.recalculate_inclination_and_desnivel()

    def update_layers(self):
        # Obter o nome da camada de polígono selecionada
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        if not selected_layer:
            # self.mostrar_mensagem("Camada de polígono não encontrada.", "Erro")
            return

        # Definir os nomes das camadas de pontos e linhas baseados na camada de polígono
        point_layer_name = f"{selected_layer.name()}_Vértices"
        line_layer_name = f"{selected_layer.name()}_Segmentos"

        # Obter as camadas pelo nome dinâmico
        point_layer = QgsProject.instance().mapLayersByName(point_layer_name)
        line_layer = QgsProject.instance().mapLayersByName(line_layer_name)

        if not point_layer or not line_layer:
            self.mostrar_mensagem(f"Camadas '{point_layer_name}' ou '{line_layer_name}' não encontradas.", "Erro")
            return

        point_layer = point_layer[0]  # Espera-se que só haja uma camada correspondente
        line_layer = line_layer[0]

        # Atualizar NovoZ na camada de pontos
        point_layer.startEditing()
        for i, feature in enumerate(point_layer.getFeatures()):
            feature['NovoZ'] = self.points[i]['NovoZ']
            point_layer.updateFeature(feature)
        point_layer.commitChanges()

        # Atualizar Inclinação na camada de linhas
        line_layer.startEditing()

        num_points = len(self.points)

        for i, feature in enumerate(line_layer.getFeatures()):
            # Conectar o último ponto ao primeiro
            point1 = self.points[i]
            point2 = self.points[(i + 1) % num_points]  # Usa o operador % para fechar o polígono

            # Recalcular a inclinação entre os pontos
            z1 = point1['NovoZ']
            z2 = point2['NovoZ']
            comprimento = feature.geometry().length()
            inclinacao = round((float(z2) - float(z1)) / comprimento if comprimento != 0 else 0, 3)

            # Atualizar o campo 'Inclinação'
            feature['Inclinação'] = inclinacao
            line_layer.updateFeature(feature)

        line_layer.commitChanges()

        self.mostrar_mensagem("Camadas de pontos e linhas atualizadas com sucesso.", "Sucesso")

    def calculate_and_create_segment_layers(self):
        # Primeiro, atualize as camadas de pontos e linhas existentes
        self.update_layers()

        # Em seguida, crie as camadas de pontos para cada segmento
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        
        if not selected_layer:
            # self.mostrar_mensagem("Camada de polígono não encontrada.", "Erro")
            return
        
        # Obter a camada de pontos
        point_layer_name = f"{selected_layer.name()}_Vértices"
        point_layer = QgsProject.instance().mapLayersByName(point_layer_name)
        
        if not point_layer:
            self.mostrar_mensagem(f"Camada '{point_layer_name}' não encontrada.", "Erro")
            return

        point_layer = point_layer[0]

        # Obter os pontos da camada de pontos
        points = [f for f in point_layer.getFeatures()]
        
        # Criar camadas de pontos para cada segmento
        self.create_point_layer_for_segments(points)

        # Atualizar o listWidgetSegmentos após criar as camadas de segmentos
        self.update_list_widget_segmentos()

        self.mostrar_mensagem("Camadas de segmentos criadas com sucesso.", "Sucesso")

    def create_point_layer_for_segments(self, points):
        # Obter o nome da camada de polígono selecionada
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        if not selected_layer:
            # self.mostrar_mensagem("Camada de polígono não encontrada.", "Erro")
            return

        # Cria ou acessa o grupo baseado no nome da camada de polígono
        root = QgsProject.instance().layerTreeRoot()
        group = root.findGroup("Segmento")
        if not group:
            group = root.addGroup("Segmento")
        else:
            # Limpar camadas existentes no grupo antes de adicionar novas camadas
            group.removeAllChildren()

        # Iterar sobre os segmentos dos pontos
        num_points = len(points)
        for i in range(num_points):
            # Definir o nome da camada de pontos para o segmento
            segment_id = f"Segmento_{i+1}_{(i+2) if i+1 < num_points else 1}"
            point_layer_name = f"{selected_layer.name()}_{segment_id}"
            point_layer = QgsVectorLayer('Point?crs=' + selected_layer.crs().authid(), point_layer_name, 'memory')
            pr = point_layer.dataProvider()

            # Adicionar campos "ID", "X", "Y", "Z", "NovoZ"
            pr.addAttributes([QgsField("ID", QVariant.Int),
                              QgsField("X", QVariant.Double),
                              QgsField("Y", QVariant.Double),
                              QgsField("Z", QVariant.Double),
                              QgsField("NovoZ", QVariant.Double)])
            point_layer.updateFields()

            # Adicionar os dois pontos do segmento à nova camada de pontos
            point1 = points[i]
            point2 = points[(i + 1) % num_points]  # Conecta o último ponto ao primeiro

            for j, point in enumerate([point1, point2]):
                feature = QgsFeature(point_layer.fields())
                x = point.geometry().asPoint().x()
                y = point.geometry().asPoint().y()

                # Definir valores para os atributos
                feature.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(x, y)))
                feature.setAttribute("ID", j + 1)
                feature.setAttribute("X", round(x, 3))
                feature.setAttribute("Y", round(y, 3))
                feature.setAttribute("Z", point['Z'])
                feature.setAttribute("NovoZ", point['NovoZ'])

                pr.addFeature(feature)

            # Adicionar a nova camada de pontos ao projeto
            QgsProject.instance().addMapLayer(point_layer, False)  # Adiciona a camada sem mostrar na árvore de camadas
            group.insertLayer(0, point_layer)  # Adiciona a camada ao grupo
            self.mostrar_mensagem(f"Camada de pontos '{point_layer_name}' criada com sucesso.", "Sucesso")

    def calculate_distance_and_update_spinbox(self):
        # Obter a camada selecionada no listWidgetSegmentos
        selected_item = self.listWidgetSegmentos.currentItem()
        if not selected_item:
            self.mostrar_mensagem("Nenhuma camada selecionada no listWidgetSegmentos.", "Erro")
            return

        # Obter a camada de pontos a partir do nome da camada
        layer_name = selected_item.text()
        layer = QgsProject.instance().mapLayersByName(layer_name)
        if not layer:
            self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
            return
        layer = layer[0]

        # Verificar se a camada possui exatamente dois pontos
        points = [f for f in layer.getFeatures()]
        if len(points) != 2:
            self.mostrar_mensagem("A camada selecionada deve conter exatamente dois pontos.", "Erro")
            return

        # Calcular a distância entre os dois pontos
        point1 = points[0].geometry().asPoint()
        point2 = points[1].geometry().asPoint()
        distance = point1.distance(point2)

        # Atualizar o valor máximo do doubleSpinBoxEquidistante
        self.doubleSpinBoxEquidistante.setMinimum(0.1)
        self.doubleSpinBoxEquidistante.setMaximum(distance)
        self.mostrar_mensagem(f"Distância entre os pontos: {distance:.3f}", "Sucesso")

    def calculate_intermediate_points(self):

        # Armazena o índice atual da seleção e define o próximo índice (se houver)
        current_row = self.listWidgetSegmentos.currentRow()
        if current_row != -1:
            self._next_segment_index = current_row + 1
        else:
            self._next_segment_index = None

        # Obter a camada selecionada no listWidgetSegmentos
        selected_item = self.listWidgetSegmentos.currentItem()
        if not selected_item:
            self.mostrar_mensagem("Nenhuma camada selecionada no listWidgetSegmentos.", "Erro")
            return

        # Obter a camada de pontos a partir do nome da camada
        layer_name = selected_item.text()
        layer = QgsProject.instance().mapLayersByName(layer_name)
        if not layer:
            self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
            return
        layer = layer[0]

        # Verificar se a camada possui exatamente dois pontos
        points = [f for f in layer.getFeatures()]
        if len(points) != 2:
            self.mostrar_mensagem("A camada selecionada deve conter exatamente dois pontos.", "Erro")
            return

        # Obter o valor selecionado no doubleSpinBoxEquidistante
        equidistance = self.doubleSpinBoxEquidistante.value()

        # Verificar se o valor da equidistância é maior que zero
        if equidistance <= 0:
            self.mostrar_mensagem("O valor da equidistância deve ser maior que zero.", "Erro")
            return

        # Calcular a distância entre os dois pontos
        point1 = points[0].geometry().asPoint()
        point2 = points[1].geometry().asPoint()
        total_distance = point1.distance(point2)

        # Verificar se a distância total entre os pontos é maior que zero
        if total_distance == 0:
            self.mostrar_mensagem("Os dois pontos são coincidentes. A distância entre eles é zero.", "Erro")
            return

        # Calcular quantos pontos intermediários podem ser gerados
        num_points = int(total_distance // equidistance)

        if num_points == 0:
            self.mostrar_mensagem("Nenhum ponto intermediário pode ser criado com o valor atual.", "Erro")
            return

        # Obter o raster selecionado para capturar o valor de Z
        selected_raster_id = self.comboBoxRaster.currentData()
        raster_layer = QgsProject.instance().mapLayer(selected_raster_id)
        if not raster_layer or not isinstance(raster_layer, QgsRasterLayer):
            self.mostrar_mensagem("Camada raster inválida ou não selecionada.", "Erro")
            return

        # Obter o NovoZ dos dois pontos
        novo_z1 = points[0]['NovoZ']
        novo_z2 = points[1]['NovoZ']

        # Calcular a inclinação entre os dois pontos
        inclination = (novo_z2 - novo_z1) / total_distance if total_distance != 0 else 0

        # Gerar o nome da nova camada com base no nome da camada de segmento
        base_name = layer_name.split('_Segmento')[0]  # Obtém o nome da camada original antes de '_Segmento'
        segment_part = layer_name.split('_Segmento_')[1]  # Obtém a parte do nome que indica o segmento (por exemplo, '1_2')
        new_layer_name = f"{base_name}_Estacas_{segment_part}"

        # Criar ou acessar o grupo "Estaqueamento"
        root = QgsProject.instance().layerTreeRoot()
        group = root.findGroup("Estaqueamento")
        if not group:
            group = root.addGroup("Estaqueamento")

        # Verificar se já existe uma camada com o mesmo nome que será criada (somente sobescreve se for a mesma camada)
        existing_layers = [l for l in QgsProject.instance().mapLayersByName(new_layer_name) if l.name() == new_layer_name]

        if existing_layers:
            # Remover a camada existente com o mesmo nome, se encontrada
            QgsProject.instance().removeMapLayer(existing_layers[0].id())

        # Criar uma nova camada de pontos para armazenar os pontos intermediários e os dois pontos externos
        crs = layer.crs()
        new_point_layer = QgsVectorLayer(f"Point?crs={crs.authid()}", new_layer_name, "memory")
        pr = new_point_layer.dataProvider()
        pr.addAttributes([QgsField("ID", QVariant.Int),
                          QgsField("X", QVariant.Double),
                          QgsField("Y", QVariant.Double),
                          QgsField("Z", QVariant.Double),
                          QgsField("NovoZ", QVariant.Double),
                          QgsField("Desnível", QVariant.Double),
                          QgsField("Acumula_dist", QVariant.Double)])  # Adiciona o campo Acumula_dist
        new_point_layer.updateFields()

        # Inicia a distância acumulada
        acumula_dist = 0

        # Adicionar o ponto inicial (ponto 1)
        feature1 = QgsFeature(new_point_layer.fields())
        feature1.setGeometry(QgsGeometry.fromPointXY(point1))
        feature1.setAttribute("ID", 1)
        feature1.setAttribute("X", round(point1.x(), 3))
        feature1.setAttribute("Y", round(point1.y(), 3))
        z_value1 = self.get_z_value(point1.x(), point1.y())  # Obter o Z da camada raster
        feature1.setAttribute("Z", z_value1 if z_value1 is not None else 0.0)
        feature1.setAttribute("NovoZ", novo_z1)
        desnivel1 = novo_z1 - z_value1 if z_value1 is not None else 0.0
        feature1.setAttribute("Desnível", round(desnivel1, 3))
        feature1.setAttribute("Acumula_dist", acumula_dist)  # Define a distância acumulada como 0 para o primeiro ponto
        pr.addFeature(feature1)

        # Calcular e adicionar os pontos intermediários
        for i in range(1, num_points + 1):
            t = i * equidistance / total_distance
            intermediate_x = (1 - t) * point1.x() + t * point2.x()
            intermediate_y = (1 - t) * point1.y() + t * point2.y()

            # Calcular o NovoZ para os pontos intermediários usando a inclinação
            novo_z_intermediate = novo_z1 + inclination * (i * equidistance)

            z_value_intermediate = self.get_z_value(intermediate_x, intermediate_y)  # Obter o Z da camada raster
            desnivel_intermediate = novo_z_intermediate - z_value_intermediate if z_value_intermediate is not None else 0.0

            # Atualizar a distância acumulada
            acumula_dist += equidistance

            # Criar a feição do ponto intermediário
            feature_intermediate = QgsFeature(new_point_layer.fields())
            feature_intermediate.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(intermediate_x, intermediate_y)))
            feature_intermediate.setAttribute("ID", i + 1)
            feature_intermediate.setAttribute("X", round(intermediate_x, 3))
            feature_intermediate.setAttribute("Y", round(intermediate_y, 3))
            feature_intermediate.setAttribute("Z", z_value_intermediate if z_value_intermediate is not None else 0.0)
            feature_intermediate.setAttribute("NovoZ", round(novo_z_intermediate, 3))
            feature_intermediate.setAttribute("Desnível", round(desnivel_intermediate, 3))
            feature_intermediate.setAttribute("Acumula_dist", acumula_dist)  # Adiciona a distância acumulada

            pr.addFeature(feature_intermediate)

        # Adicionar o ponto final (ponto 2)
        acumula_dist = total_distance  # Atualiza corretamente a distância acumulada no ponto final
        feature2 = QgsFeature(new_point_layer.fields())
        feature2.setGeometry(QgsGeometry.fromPointXY(point2))
        feature2.setAttribute("ID", num_points + 2)
        feature2.setAttribute("X", round(point2.x(), 3))
        feature2.setAttribute("Y", round(point2.y(), 3))
        z_value2 = self.get_z_value(point2.x(), point2.y())  # Obter o Z da camada raster
        feature2.setAttribute("Z", z_value2 if z_value2 is not None else 0.0)
        feature2.setAttribute("NovoZ", novo_z2)
        desnivel2 = novo_z2 - z_value2 if z_value2 is not None else 0.0
        feature2.setAttribute("Desnível", round(desnivel2, 3))
        feature2.setAttribute("Acumula_dist", acumula_dist)  # Adiciona a distância acumulada correta para o último ponto
        pr.addFeature(feature2)

        # Adicionar a nova camada de pontos ao grupo "Estaqueamento"
        QgsProject.instance().addMapLayer(new_point_layer, False)  # Adiciona a camada sem mostrar na árvore de camadas
        group.insertLayer(0, new_point_layer)  # Adiciona a camada ao grupo

        self.mostrar_mensagem(f"Camada '{new_layer_name}' criada e adicionada ao grupo 'Estaqueamento'.", "Sucesso")

        # Aplicar os rótulos à nova camada
        self.set_label_for_layer(new_point_layer, "Desnível")

        # Atualizar o listWidgetEstacas logo após a criação da camada
        self.update_list_widget_estaqueamentos()

    def set_label_for_layer(self, layer, field_name):
        """
        Configura o rótulo para uma camada no QGIS, com base em um campo específico e formatações adicionais.
        """
        label_settings = QgsPalLayerSettings()
        label_settings.drawBackground = True  # Ativa o fundo do rótulo
        label_settings.fieldName = field_name  # Define o campo a ser rotulado

        # Configura a formatação do texto
        text_format = label_settings.format()
        font = text_format.font()
        font.setItalic(True)
        font.setBold(True)
        text_format.setFont(font)

        # Configura a cor do rótulo com base em uma expressão condicional
        color_expression = """CASE
                                WHEN "Desnível" < 0 THEN '255,0,0'  -- Vermelho
                                WHEN "Desnível" > 0 THEN '0,0,255'  -- Azul
                                ELSE '0,0,0'  -- Preto
                              END"""

        properties = label_settings.dataDefinedProperties()
        properties.setProperty(QgsPalLayerSettings.Color, QgsProperty.fromExpression(color_expression))

        # Defina o tamanho da fonte usando propriedades definidas por dados
        properties.setProperty(QgsPalLayerSettings.Size, QgsProperty.fromValue(8))
        label_settings.setDataDefinedProperties(properties)

        # Configura o fundo branco para os rótulos
        background_color = text_format.background()
        background_color.setEnabled(True)
        background_color.setFillColor(QColor(255, 255, 255))
        text_format.setBackground(background_color)

        label_settings.setFormat(text_format)

        # Ativa o rótulo para a camada
        layer.setLabelsEnabled(True)
        layer.setLabeling(QgsVectorLayerSimpleLabeling(label_settings))
        layer.triggerRepaint()

    def get_z_value(self, x, y):
        selected_raster_id = self.comboBoxRaster.currentData()
        raster_layer = QgsProject.instance().mapLayer(selected_raster_id)
        
        if not raster_layer or not isinstance(raster_layer, QgsRasterLayer):
            self.mostrar_mensagem("Camada raster inválida ou não selecionada.", "Erro")
            return None
        
        identify_result = raster_layer.dataProvider().identify(QgsPointXY(x, y), QgsRaster.IdentifyFormatValue)
        
        if identify_result.isValid():
            results = identify_result.results()
            if results:
                band_key = list(results.keys())[0]
                z_value = results.get(band_key)
                if z_value is not None:
                    return round(float(z_value), 3)

        # Se não conseguir identificar o valor Z, retorna None
        self.mostrar_mensagem(f"Não foi possível identificar o valor Z nas coordenadas ({x}, {y}).", "Erro")
        return None

    def calculate_talude(self):

        # Armazena o índice atual do listWidgetEstacas e define o próximo índice, se existir
        current_row = self.listWidgetEstacas.currentRow()
        if current_row != -1:
            self._next_estaca_index = current_row + 1
        else:
            self._next_estaca_index = None

        # Obter a camada selecionada no listWidgetEstacas
        selected_item = self.listWidgetEstacas.currentItem()
        if not selected_item:
            self.mostrar_mensagem("Nenhuma camada selecionada no listWidgetEstacas.", "Erro")
            return

        # Obter a camada de pontos a partir do nome da camada
        layer_name = selected_item.text()
        layer = QgsProject.instance().mapLayersByName(layer_name)
        if not layer:
            self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
            return
        layer = layer[0]

        # Verificar se a camada contém pelo menos dois pontos
        points = [f for f in layer.getFeatures()]
        if len(points) < 2:
            self.mostrar_mensagem("A camada selecionada deve conter pelo menos dois pontos.", "Erro")
            return

        # Obter o raster selecionado para capturar o valor de Z
        selected_raster_id = self.comboBoxRaster.currentData()
        raster_layer = QgsProject.instance().mapLayer(selected_raster_id)
        if not raster_layer or not isinstance(raster_layer, QgsRasterLayer):
            self.mostrar_mensagem("Camada raster inválida ou não selecionada.", "Erro")
            return

        # Verificar qual radio button está selecionado
        if self.radioButtonEsquerda.isChecked():
            lado = "esquerda"
            direction_factor = 1  # Para criar o talude à esquerda
        elif self.radioButtonDireita.isChecked():
            lado = "direita"
            direction_factor = -1  # Para criar o talude à direita
        else:
            self.mostrar_mensagem("Selecione se o talude será criado à esquerda ou à direita.", "Erro")
            return

        # Nome da nova camada de talude
        talude_layer_name = f"{layer_name}_Talude_{lado.capitalize()}"

        # Verificar se já existe uma camada com o mesmo nome
        existing_layers = QgsProject.instance().mapLayersByName(talude_layer_name)
        if existing_layers:
            QgsProject.instance().removeMapLayer(existing_layers[0].id())

        # Criar a nova camada de pontos para o talude
        crs = layer.crs()
        talude_layer = QgsVectorLayer(f"Point?crs={crs.authid()}", talude_layer_name, "memory")
        pr = talude_layer.dataProvider()
        pr.addAttributes([QgsField("ID", QVariant.Int),
                          QgsField("X", QVariant.Double),
                          QgsField("Y", QVariant.Double),
                          QgsField("Z", QVariant.Double),
                          QgsField("NovoZ", QVariant.Double),
                          QgsField("Desnível", QVariant.Double)])
        talude_layer.updateFields()

        # Percorrer os pontos e calcular a posição dos pontos de talude
        num_points = len(points)
        for i in range(num_points):
            # Para o último ponto, conecte ao primeiro ponto
            point1 = points[i].geometry().asPoint()
            if i == num_points - 1:
                point2 = points[0].geometry().asPoint()  # Conecta ao primeiro ponto para fechar o polígono
            else:
                point2 = points[i + 1].geometry().asPoint()

            # Calcular o vetor perpendicular com base na função simplificada
            dx = point2.x() - point1.x()
            dy = point2.y() - point1.y()
            
            # Normalizar o vetor
            length = math.sqrt(dx ** 2 + dy ** 2)
            if length == 0:
                continue
            
            # Perpendicular unitário
            perp_dx = -dy / length
            perp_dy = dx / length

            # Usar o "Desnível" para definir a altura do talude
            talude_height = points[i]['Desnível']

            # Calcular as novas coordenadas do talude
            talude_x = point1.x() + direction_factor * abs(talude_height) * perp_dx
            talude_y = point1.y() + direction_factor * abs(talude_height) * perp_dy

            # **Ajuste aqui para o último ponto**: Verificar se o último ponto está sendo invertido
            if i == num_points - 1:
                talude_x = point1.x() + direction_factor * abs(talude_height) * (-perp_dx)
                talude_y = point1.y() + direction_factor * abs(talude_height) * (-perp_dy)

            # Obter o Z do raster na posição do talude
            talude_z = self.get_z_value(talude_x, talude_y)
            if talude_z is None:
                talude_z = 0.0

            # Obter o "NovoZ" do ponto atual para calcular o desnível
            novo_z1 = points[i]["NovoZ"]
            desnivel_talude = novo_z1 - talude_z

            # Criar a feição para o ponto de talude
            talude_feature = QgsFeature(talude_layer.fields())
            talude_feature.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(talude_x, talude_y)))
            talude_feature.setAttribute("ID", i + 1)
            talude_feature.setAttribute("X", round(talude_x, 3))
            talude_feature.setAttribute("Y", round(talude_y, 3))
            talude_feature.setAttribute("Z", talude_z)
            talude_feature.setAttribute("NovoZ", round(novo_z1, 3))
            talude_feature.setAttribute("Desnível", round(desnivel_talude, 3))

            pr.addFeature(talude_feature)

        # Adicionar a camada de talude ao projeto dentro do grupo "Talude"
        root = QgsProject.instance().layerTreeRoot()
        group = root.findGroup("Talude")
        if not group:
            group = root.addGroup("Talude")

        QgsProject.instance().addMapLayer(talude_layer, False)
        group.insertLayer(0, talude_layer)

        self.mostrar_mensagem(f"Camada '{talude_layer_name}' de talude criada com sucesso.", "Sucesso")

        # Atualizar o listWidgetTalude logo após a criação da camada
        self.update_list_widget_talude()

    def find_matching_talude_layers(self, estacas_layer_name, talude_layers_dict):
        """Encontra as camadas de Talude correspondentes com base no nome da camada de Estacas."""
        matching_layers = []
        for talude_layer_name in talude_layers_dict.keys():
            # Verificar se o nome da camada de Talude contém o nome da camada de Estacas
            if estacas_layer_name in talude_layer_name:
                matching_layers.append(talude_layer_name)
        return matching_layers

    def create_connections_between_layers(self, estacas_points, talude_points, crs, base_layer_name, line_layer, feature_id_start):
        """Cria conexões entre os pontos das camadas de Estacas e Talude correspondentes e adiciona à camada de linhas única."""
        if not estacas_points or not talude_points:
            self.mostrar_mensagem(f"Não há pontos suficientes em Estacas ou Talude para criar conexões para '{base_layer_name}'.", "Erro")
            return feature_id_start

        pr_line = line_layer.dataProvider()
        feature_id = feature_id_start  # Atualiza o ID com base no valor passado

        line_features = []

        # Ordenar os pontos se necessário
        estacas_points_sorted = sorted(estacas_points, key=lambda x: x[0]['ID'])
        talude_points_sorted = sorted(talude_points, key=lambda x: x[0]['ID'])

        # Coletar os pontos para formar o contorno externo
        contour_points = []

        # Reprojetar e coletar pontos de Estacas
        for estaca_feature, estaca_geom, estaca_crs in estacas_points_sorted:
            if estaca_crs != crs:
                transform = QgsCoordinateTransform(estaca_crs, crs, QgsProject.instance())
                estaca_geom = estaca_geom.clone()
                estaca_geom.transform(transform)
            estaca_point = estaca_geom.asPoint()
            contour_points.append(estaca_point)

        # Reprojetar e coletar pontos de Talude em ordem reversa
        for talude_feature, talude_geom, talude_crs in reversed(talude_points_sorted):
            if talude_crs != crs:
                transform = QgsCoordinateTransform(talude_crs, crs, QgsProject.instance())
                talude_geom = talude_geom.clone()
                talude_geom.transform(transform)
            talude_point = talude_geom.asPoint()
            contour_points.append(talude_point)

        # Fechar o contorno conectando o último ponto ao primeiro, se necessário
        if contour_points[0] != contour_points[-1]:
            contour_points.append(contour_points[0])

        # Criar a geometria da linha do contorno
        contour_geom = QgsGeometry.fromPolylineXY(contour_points)

        # Criar a feição de linha para o contorno
        contour_feature = QgsFeature(line_layer.fields())
        contour_feature.setGeometry(contour_geom)
        contour_feature.setAttribute('ID', feature_id)
        contour_feature.setAttribute('Tipo', 'Contorno')
        contour_feature.setAttribute('Nome', base_layer_name)
        line_features.append(contour_feature)
        feature_id += 1

        # Criar as linhas inclinadas conectando os pontos correspondentes de Estacas e Talude
        min_length = min(len(estacas_points_sorted), len(talude_points_sorted))
        for i in range(min_length):
            estaca_feature, estaca_geom, _ = estacas_points_sorted[i]
            talude_feature, talude_geom, _ = talude_points_sorted[i]

            estaca_point = estaca_geom.asPoint()
            talude_point = talude_geom.asPoint()

            # Criar a geometria da linha inclinada
            inclined_line_geom = QgsGeometry.fromPolylineXY([estaca_point, talude_point])

            # Criar a feição de linha inclinada
            inclined_line_feature = QgsFeature(line_layer.fields())
            inclined_line_feature.setGeometry(inclined_line_geom)
            inclined_line_feature.setAttribute('ID', feature_id)
            inclined_line_feature.setAttribute('Tipo', 'Ligação')
            inclined_line_feature.setAttribute('Nome', base_layer_name)
            line_features.append(inclined_line_feature)
            feature_id += 1

        # Adicionar as feições à camada de linhas única
        pr_line.addFeatures(line_features)

        # Retornar o próximo ID disponível
        return feature_id

    def process_matching_layers(self, estacas_layer, talude_layers, crs, line_layer, pr_line, feature_id):
        """Processa as camadas de Estacas e Talude correspondentes e cria as conexões."""
        estacas_points = []
        # Coletar os pontos da camada de Estacas
        for feature in estacas_layer.getFeatures():
            geom = feature.geometry()
            if geom is None or geom.isEmpty():
                continue
            if geom.type() != QgsWkbTypes.PointGeometry:
                self.mostrar_mensagem(f"A camada '{estacas_layer.name()}' não é uma camada de pontos.", "Erro")
                continue
            estacas_points.append((feature, geom, estacas_layer.crs()))

        # Coletar os pontos das camadas de Talude correspondentes
        talude_points = []
        for talude_layer in talude_layers:
            for feature in talude_layer.getFeatures():
                geom = feature.geometry()
                if geom is None or geom.isEmpty():
                    continue
                if geom.type() != QgsWkbTypes.PointGeometry:
                    self.mostrar_mensagem(f"A camada '{talude_layer.name()}' não é uma camada de pontos.", "Erro")
                    continue
                talude_points.append((feature, geom, talude_layer.crs()))

        # Criar as conexões e adicionar as feições à camada de linhas
        feature_id = self.create_connections_between_layers(estacas_points, talude_points, crs, estacas_layer.name(), line_layer, pr_line, feature_id)

        return feature_id

    def check_layers_in_group_and_update(self, removed_layers_ids):
        # Obter o nome da camada de polígono selecionada
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)

        if not selected_layer:
            # self.mostrar_mensagem("Camada de polígono não encontrada.", "Erro")
            self.clear_all_list_widgets()
            return

        # Verifica os grupos e atualiza os listWidgets
        self.check_group_and_update("Segmento", self.listWidgetSegmentos, removed_layers_ids)
        self.check_group_and_update("Estaqueamento", self.listWidgetEstacas, removed_layers_ids)
        self.check_group_and_update("Talude", self.listWidgetTalude, removed_layers_ids)

    def check_group_and_update(self, group_name, list_widget, removed_layers_ids):
        root = QgsProject.instance().layerTreeRoot()
        group = root.findGroup(group_name)

        if not group:
            # self.mostrar_mensagem(f"O grupo '{group_name}' foi removido.", "Erro")
            list_widget.clear()
        else:
            group_layers = [layer.layerId() for layer in group.findLayers()]
            for layer_id in removed_layers_ids:
                if layer_id in group_layers:
                    self.mostrar_mensagem(f"Uma camada foi removida do grupo '{group_name}'.", "Erro")
                    self.update_list_widget(list_widget, group_name, group_name.capitalize())
                    break

    def clear_all_list_widgets(self):
        self.listWidgetSegmentos.clear()
        self.listWidgetEstacas.clear()
        self.listWidgetTalude.clear()

    def update_list_widget_talude(self):
        self.update_list_widget(self.listWidgetTalude, "Talude", "Taludes")

        # Chama a função para atualizar o estado do pushButtonJuntar
        self.update_pushButtonJuntar_state()

    def update_list_widget_estaqueamentos(self):
        self.update_list_widget(self.listWidgetEstacas, "Estaqueamento", "Estaqueamentos")

        # Chama a função para atualizar o estado dos botões de rádio
        self.update_radio_buttons_state()

        # Chama a função para atualizar o estado do pushButtonJuntar
        self.update_pushButtonJuntar_state()

    def update_list_widget_segmentos(self):
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        
        if not selected_layer:
            # self.mostrar_mensagem("Camada de polígono não encontrada.", "Erro")
            return

        self.update_list_widget(self.listWidgetSegmentos, "Segmento", "Segmentos")

    def update_list_widget(self, list_widget, group_name, header_text):
        # Limpa o listWidget antes de adicionar novas entradas
        list_widget.clear()

        # Acessa o grupo específico no projeto
        root = QgsProject.instance().layerTreeRoot()
        group = root.findGroup(group_name)

        if not group:
            # self.mostrar_mensagem(f"O grupo '{group_name}' não foi encontrado.", "Erro")
            return  # Sai da função se o grupo não for encontrado

        # Adiciona o cabeçalho com estilo
        header_label = QLabel(header_text)
        header_label.setAlignment(Qt.AlignCenter)
        header_label.setFont(QFont("Arial", 8, QFont.Bold))

        # Estiliza a borda para criar um efeito 3D
        header_label.setStyleSheet("""
            QLabel {
                background-color: #C0C0C0;  /* Cinza 3D */
                color: #000000;  /* Cor do texto */
                padding: 2px;
                border: 1px solid #A0A0A0;  /* Borda 3D mais clara */
                border-left-color: #E0E0E0;  /* Borda 3D mais clara à esquerda */
                border-top-color: #E0E0E0;  /* Borda 3D mais clara no topo */
                border-right-color: #808080; /* Borda 3D mais escura à direita */
                border-bottom-color: #808080; /* Borda 3D mais escura na parte inferior */
            }
        """)

        header_item = QListWidgetItem(list_widget)
        header_item.setFlags(Qt.NoItemFlags)  # Desabilita a seleção do cabeçalho
        list_widget.addItem(header_item)
        list_widget.setItemWidget(header_item, header_label)
        header_item.setSizeHint(header_label.sizeHint())

        # Estilo para os itens da lista
        list_widget.setStyleSheet("""
            QListWidget::item:hover {
                background-color: #aaffff;  /* Azul claro no hover */
            }
            QListWidget::item:selected {
                background-color: #0055ff;  /* Azul claro para itens selecionados */
                color: black;
            }
            QListWidget::item:hover:selected {
                background-color: #0055ff;  /* Mantém a seleção ao passar o mouse sobre o item selecionado */
            }
            QListWidget::item {
                padding: 1px;
            }
        """)

        for layer in group.findLayers():
            layer_obj = layer.layer()
            if isinstance(layer_obj, QgsVectorLayer) and layer_obj.geometryType() == QgsWkbTypes.PointGeometry:
                item = QListWidgetItem(layer_obj.name())
                item.setData(Qt.UserRole, layer_obj.id())  # Associa o ID da camada ao item
                item.setToolTip(layer_obj.name())  # Define o toolTip com o nome da camada
                list_widget.addItem(item)
#===========================================================
        # Se estivermos atualizando o listWidgetSegmentos e houver uma seleção "guardada", define a nova seleção
        if list_widget == self.listWidgetSegmentos and hasattr(self, "_next_segment_index") and self._next_segment_index is not None:
            total = list_widget.count()
            # Se houver mais itens do que o índice desejado, seleciona-o (caso contrário, limpa a seleção)
            if self._next_segment_index < total:
                list_widget.setCurrentRow(self._next_segment_index)
                self._log_message(f"Seleção movida para o item {self._next_segment_index}.", Qgis.Info)
            else:
                list_widget.setCurrentRow(-1)
                self._log_message("Não há próximo item: seleção removida.", Qgis.Info)
            self._next_segment_index = None

        elif list_widget == self.listWidgetEstacas and hasattr(self, "_next_estaca_index") and self._next_estaca_index is not None:
            total = list_widget.count()
            if self._next_estaca_index < total:
                list_widget.setCurrentRow(self._next_estaca_index)
                self._log_message(f"Seleção movida para o item {self._next_estaca_index} em Estaqueamentos.", Qgis.Info)
            else:
                list_widget.setCurrentRow(-1)
                self._log_message("Não há próximo item em Estaqueamentos: seleção removida.", Qgis.Info)
            self._next_estaca_index = None

    def check_estaqueamento_button_state(self):
        """Verifica se o botão Estaqueamento deve ser ativado."""
        # Verifica se há um item selecionado no listWidgetSegmentos e se o valor do doubleSpinBoxEquidistante é maior que 0
        has_selection = self.listWidgetSegmentos.selectedItems()
        equidistance_value = self.doubleSpinBoxEquidistante.value()

        if has_selection and equidistance_value > 0:
            self.pushButtonEstaqueamento.setEnabled(True)
        else:
            self.pushButtonEstaqueamento.setEnabled(False)

    def update_doubleSpinBoxEquidistante_state(self):
        """
        Ativa ou desativa o doubleSpinBoxEquidistante com base na seleção no listWidgetSegmentos
        e ajusta o valor máximo do doubleSpinBoxEquidistante para ser o comprimento da linha selecionada.
        """
        # Verificar se há itens no listWidgetSegmentos
        if self.listWidgetSegmentos.count() == 0:
            self.doubleSpinBoxEquidistante.setEnabled(False)  # Desativa o doubleSpinBoxEquidistante se não houver itens
            self.doubleSpinBoxEquidistante.setToolTip("Nenhuma camada de segmento disponível.")
            return

        selected_item = self.listWidgetSegmentos.currentItem()

        # Verificar se há uma camada selecionada no listWidgetSegmentos
        if not selected_item:
            self.doubleSpinBoxEquidistante.setEnabled(False)  # Desativa o doubleSpinBoxEquidistante se não houver seleção
            self.doubleSpinBoxEquidistante.setToolTip("Selecione uma camada de segmento primeiro.")
            return

        # Obter a camada de pontos a partir do nome da camada
        layer_name = selected_item.text()
        layer = QgsProject.instance().mapLayersByName(layer_name)
        if not layer:
            self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
            return
        layer = layer[0]

        # Verificar se a camada possui exatamente dois pontos
        points = [f for f in layer.getFeatures()]
        if len(points) != 2:
            self.mostrar_mensagem("A camada selecionada deve conter exatamente dois pontos.", "Erro")
            return

        # Calcular a distância entre os dois pontos
        point1 = points[0].geometry().asPoint()
        point2 = points[1].geometry().asPoint()
        total_distance = point1.distance(point2)

        # Ativar o doubleSpinBoxEquidistante e definir o valor máximo
        self.doubleSpinBoxEquidistante.setEnabled(True)
        self.doubleSpinBoxEquidistante.setMaximum(total_distance)
        self.doubleSpinBoxEquidistante.setToolTip(f"Distância total: {total_distance:.2f} m. Selecione uma equidistância menor ou igual.")

        # Atualizar o valor do toolTip para mostrar a quantidade de pontos que será criada
        self.calculate_and_update_tooltip()

    def calculate_and_update_tooltip(self):
        """
        Calcula a quantidade de pontos que será criada com base no valor da equidistância
        e na linha selecionada no listWidgetSegmentos, e atualiza o toolTip.
        """
        selected_item = self.listWidgetSegmentos.currentItem()

        # Verificar se há uma camada selecionada no listWidgetSegmentos
        if not selected_item:
            self.doubleSpinBoxEquidistante.setToolTip("Nenhuma camada de segmento selecionada.")  # Atualizar o toolTip
            return

        # Obter a camada de pontos a partir do nome da camada
        layer_name = selected_item.text()
        layer = QgsProject.instance().mapLayersByName(layer_name)
        if not layer:
            self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
            return
        layer = layer[0]

        # Verificar se a camada possui exatamente dois pontos
        points = [f for f in layer.getFeatures()]
        if len(points) != 2:
            self.mostrar_mensagem("A camada selecionada deve conter exatamente dois pontos.", "Erro")
            return

        # Calcular a distância entre os dois pontos
        point1 = points[0].geometry().asPoint()
        point2 = points[1].geometry().asPoint()
        total_distance = point1.distance(point2)

        # Obter o valor de equidistância
        equidistance = self.doubleSpinBoxEquidistante.value()

        # Verificar se o valor da equidistância é válido
        if equidistance > 0 and total_distance > 0:
            # Calcular a quantidade de pontos
            num_points = int(total_distance // equidistance)
            tooltip_text = f"Distância total: {total_distance:.2f} m | Pontos a criar: {num_points}"
        else:
            tooltip_text = "Equidistância inválida ou distância zero."

        # Atualizar o toolTip do doubleSpinBox
        self.doubleSpinBoxEquidistante.setToolTip(tooltip_text)

    def update_pushButtonTalude_state(self):
        """
        Ativa ou desativa o pushButtonTalude com base na seleção no listWidgetEstacas.
        """
        selected_item = self.listWidgetEstacas.currentItem()

        # Verificar se há uma camada selecionada no listWidgetEstacas
        if not selected_item:
            self.pushButtonTalude.setEnabled(False)  # Desativa o botão se nenhuma camada estiver selecionada
            return

        # Se houver uma camada selecionada, ativa o botão
        self.pushButtonTalude.setEnabled(True)

    def update_pushButtonVertices_state(self):
        """
        Ativa ou desativa o pushButtonVertices com base nas seguintes condições:
        - Caso não haja camada de polígono no comboBoxPoligono.
        - Caso não haja camada no comboBoxRaster.
        - Caso as camadas de polígono e raster tenham georreferências diferentes.
        - Caso alguma das camadas esteja em coordenadas geográficas.
        - Se a camada de polígono tenha mais de uma feição e não haja uma única selecionada.
        - Se alguma das camadas não tiver projeção.
        """
        selected_polygon_index = self.comboBoxPoligono.currentIndex()
        selected_raster_index = self.comboBoxRaster.currentIndex()

        # Verifica se há uma camada de polígono e raster selecionada
        if selected_polygon_index == -1 or selected_raster_index == -1:
            self.pushButtonVertices.setEnabled(False)
            return

        # Obtém a camada de polígono e raster selecionada
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_raster_id = self.comboBoxRaster.currentData()
        polygon_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        raster_layer = QgsProject.instance().mapLayer(selected_raster_id)

        # Verifica se ambas as camadas são válidas
        if not polygon_layer or not raster_layer:
            self.pushButtonVertices.setEnabled(False)
            return

        # Verifica se as camadas têm projeção
        if not polygon_layer.crs().isValid() or not raster_layer.crs().isValid():
            if self.isVisible():
                self.mostrar_mensagem("Uma ou ambas as camadas não têm projeção definida.", "Erro")
            self.pushButtonVertices.setEnabled(False)
            return

        # Verifica se as camadas têm sistemas de coordenadas diferentes
        if polygon_layer.crs() != raster_layer.crs():
            if self.isVisible():
                self.mostrar_mensagem("As camadas de polígono e raster têm georreferências diferentes.", "Erro")
            self.pushButtonVertices.setEnabled(False)
            return

        # Verifica se alguma das camadas está em coordenadas geográficas (ex.: EPSG:4326)
        if polygon_layer.crs().isGeographic() or raster_layer.crs().isGeographic():
            if self.isVisible():
                self.mostrar_mensagem("Uma ou ambas as camadas estão em coordenadas geográficas.", "Erro")
            self.pushButtonVertices.setEnabled(False)
            return

        # Verifica se a camada de polígono tem mais de uma feição e se não há uma única selecionada
        if polygon_layer.featureCount() > 1 and polygon_layer.selectedFeatureCount() != 1:
            if self.isVisible():
                self.mostrar_mensagem("A camada de polígono tem mais de uma feição e nenhuma feição única está selecionada.", "Erro")
            self.pushButtonVertices.setEnabled(False)
            return

        # Se todas as condições forem atendidas, ativa o botão
        self.pushButtonVertices.setEnabled(True)

    def on_pushButtonGrafico_clicked(self):
        """Método para ser chamado quando o botão de gráfico for clicado."""
        # Obter a camada selecionada no listWidgetEstacas
        selected_item = self.listWidgetEstacas.currentItem()
        if not selected_item:
            self.mostrar_mensagem("Nenhuma camada de estacas selecionada.", "Erro")
            return

        estacas_layer_name = selected_item.text()
        estacas_layer = self.grafico_manager.find_layer(estacas_layer_name, QgsVectorLayer)
        if not estacas_layer:
            self.mostrar_mensagem(f"Camada de estacas '{estacas_layer_name}' não encontrada.", "Erro")
            return

        # Obter a camada raster selecionada
        raster_layer_id = self.comboBoxRaster.currentData()
        raster_layer = QgsProject.instance().mapLayer(raster_layer_id)
        if not raster_layer:
            self.mostrar_mensagem("Camada raster não encontrada.", "Erro")
            return

        # Criar a camada de pontos de apoio temporariamente para uso no gráfico
        pontos_apoio_layer = self.grafico_manager.create_support_points_layer(estacas_layer, raster_layer)

        # Passar o caminho do logotipo para o GraficoManager
        self.grafico_manager.logo_path = self.logo_path

        # Gerar o gráfico com base nas camadas de estacas e pontos de apoio (sem adicionar ao projeto)
        self.grafico_manager.plot_layers(estacas_layer, pontos_apoio_layer, 45, 45)

    def on_pushButtonJuntar_clicked(self):
        """Método para unir camadas dos listWidgets Estacas e Talude em uma única camada de pontos e criar conexões de linhas."""
        # Inicializa conjuntos e listas
        unique_coords = set()
        unique_features = []
        crs = None  # Variável para armazenar o CRS da camada unificada

        # Listas para armazenar os pontos de Estacas e Talude separadamente
        estacas_points_global = []
        talude_points_global = []

        # Função para processar camadas de um dado listWidget
        def process_list_widget(list_widget, cota_field_name, points_list, layer_type):
            nonlocal crs  # Permite modificar a variável 'crs' definida no escopo externo
            for index in range(list_widget.count()):
                item = list_widget.item(index)
                # Pula o cabeçalho se presente
                if not item or not item.text() or item.text() in ["Estaqueamentos", "Taludes"]:
                    continue
                layer_name = item.text()
                layer = QgsProject.instance().mapLayersByName(layer_name)
                if not layer:
                    self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
                    continue
                layer = layer[0]
                if not isinstance(layer, QgsVectorLayer):
                    continue

                # Se o CRS ainda não foi definido, define-o com o CRS da camada atual
                if crs is None:
                    crs = layer.crs()
                elif crs != layer.crs():
                    self.mostrar_mensagem(f"As camadas têm sistemas de coordenadas diferentes. Reprojetando para o CRS da primeira camada.", "Erro")
                    # Reprojetar geometrias se necessário

                for feature in layer.getFeatures():
                    geom = feature.geometry()
                    if geom is None or geom.isEmpty():
                        continue
                    if geom.type() != QgsWkbTypes.PointGeometry:
                        self.mostrar_mensagem(f"A camada '{layer_name}' não é uma camada de pontos.", "Erro")
                        continue
                    point = geom.asPoint()
                    coord_tuple = (round(point.x(), 6), round(point.y(), 6))  # Arredonda para 6 casas decimais
                    if coord_tuple not in unique_coords:
                        unique_coords.add(coord_tuple)
                        # Obter os atributos necessários
                        attributes = {
                            'ID': feature['ID'],
                            'X': feature['X'],
                            'Y': feature['Y'],
                            'Cota': feature[cota_field_name]
                        }
                        unique_features.append((attributes, geom, layer.crs()))
                    # Adicionar o ponto à lista correspondente (Estacas ou Talude)
                    points_list.append((feature, geom, layer.crs(), layer_name))

        # Processa as camadas do listWidgetEstacas (usa 'NovoZ' como 'Cota')
        process_list_widget(self.listWidgetEstacas, 'NovoZ', estacas_points_global, 'Estacas')
        # Processa as camadas do listWidgetTalude (usa 'Z' como 'Cota')
        process_list_widget(self.listWidgetTalude, 'Z', talude_points_global, 'Talude')

        if not unique_features:
            self.mostrar_mensagem("Nenhuma feição encontrada nas camadas selecionadas.", "Erro")
            return

        if crs is None or not crs.isValid():
            crs = QgsCoordinateReferenceSystem('EPSG:4326')
            self.mostrar_mensagem(f"CRS inválido. Usando EPSG:4326.", "Erro")

        # Cria uma nova camada de memória para armazenar as feições unidas
        merged_layer = QgsVectorLayer(f"Point?crs={crs.authid()}", "Camada_Junta", "memory")
        pr = merged_layer.dataProvider()

        # Define os campos: 'ID', 'X', 'Y', 'Cota'
        fields = [
            QgsField('ID', QVariant.Int),
            QgsField('X', QVariant.Double),
            QgsField('Y', QVariant.Double),
            QgsField('Cota', QVariant.Double)
        ]
        pr.addAttributes(fields)
        merged_layer.updateFields()

        # Adiciona as feições únicas à nova camada
        new_features = []
        for attributes, geom, feature_crs in unique_features:
            new_feature = QgsFeature(merged_layer.fields())
            # Reprojetar a geometria se necessário
            if feature_crs != crs:
                transform = QgsCoordinateTransform(feature_crs, crs, QgsProject.instance())
                geom = geom.clone()
                geom.transform(transform)
            new_feature.setGeometry(geom)
            # Definir os atributos
            new_feature.setAttribute('ID', attributes['ID'])
            new_feature.setAttribute('X', attributes['X'])
            new_feature.setAttribute('Y', attributes['Y'])
            new_feature.setAttribute('Cota', attributes['Cota'])
            new_features.append(new_feature)
        pr.addFeatures(new_features)

        # Adiciona a camada unificada ao projeto
        QgsProject.instance().addMapLayer(merged_layer)
        self.mostrar_mensagem(f"Camadas unidas com sucesso em '{merged_layer.name()}'.", "Sucesso")

        # Após a criação da camada unificada de pontos, criar a camada de linhas
        # Criar a camada de linhas única
        line_layer_name = "Linhas_Platô_Talude"
        line_layer = QgsVectorLayer(f"LineString?crs={crs.authid()}", line_layer_name, "memory")
        pr_line = line_layer.dataProvider()
        # Definir campos, se necessário
        line_fields = [
            QgsField('ID', QVariant.Int),
            QgsField('Tipo', QVariant.String),
            QgsField('Nome', QVariant.String)
        ]
        pr_line.addAttributes(line_fields)
        line_layer.updateFields()

        feature_id = 1  # Contador de IDs das feições na camada de linhas

        # Agrupar os pontos de Estacas e Talude por camada
        estacas_layers = {}
        for feature, geom, feature_crs, layer_name in estacas_points_global:
            if layer_name not in estacas_layers:
                estacas_layers[layer_name] = []
            estacas_layers[layer_name].append((feature, geom, feature_crs))

        talude_layers = {}
        for feature, geom, feature_crs, layer_name in talude_points_global:
            if layer_name not in talude_layers:
                talude_layers[layer_name] = []
            talude_layers[layer_name].append((feature, geom, feature_crs))

        # Processar cada camada de Estacas e encontrar a correspondente de Talude
        for estacas_layer_name, estacas_points in estacas_layers.items():
            # Encontrar a(s) camada(s) de Talude correspondente(s)
            matching_talude_layers = self.find_matching_talude_layers(estacas_layer_name, talude_layers)

            if not matching_talude_layers:
                self.mostrar_mensagem(f"Nenhuma camada de Talude correspondente encontrada para '{estacas_layer_name}'.", "Erro")
                continue

            # Coletar os pontos de Talude correspondentes
            talude_points = []
            for talude_layer_name in matching_talude_layers:
                talude_points.extend(talude_layers[talude_layer_name])

            # Criar as conexões e adicionar à camada de linhas única
            feature_id = self.create_connections_between_layers(estacas_points, talude_points, crs, estacas_layer_name, line_layer, feature_id)

        # Adicionar a camada de linhas única ao projeto após o processamento de todas as camadas
        QgsProject.instance().addMapLayer(line_layer)
        self.mostrar_mensagem(f"Camada de linhas '{line_layer_name}' criada com sucesso.", "Sucesso")

    def update_radio_buttons_state(self):
        """
        Habilita ou desabilita os botões de rádio (radioButtonEsquerda e radioButtonDireita)
        com base no conteúdo do listWidgetEstacas.
        """
        # Verifica se há itens no listWidgetEstacas
        if self.listWidgetEstacas.count() > 0:
            # Se houver itens, habilita os botões de rádio
            self.radioButtonEsquerda.setEnabled(True)
            self.radioButtonDireita.setEnabled(True)
        else:
            # Se não houver itens, desabilita os botões de rádio
            self.radioButtonEsquerda.setEnabled(False)
            self.radioButtonDireita.setEnabled(False)

    def update_pushButtonGrafico_state(self):
        """
        Habilita ou desabilita o pushButtonGrafico com base na seleção de camadas no listWidgetEstacas.
        """
        # Verifica se há uma camada selecionada no listWidgetEstacas
        if self.listWidgetEstacas.currentItem() is not None:
            # Se houver uma camada selecionada, habilita o pushButtonGrafico
            self.pushButtonGrafico.setEnabled(True)
        else:
            # Se não houver camada selecionada, desabilita o pushButtonGrafico
            self.pushButtonGrafico.setEnabled(False)

    def update_pushButtonJuntar_state(self):
        """
        Habilita ou desabilita o pushButtonJuntar com base no conteúdo do listWidgetEstacas e listWidgetTalude.
        """
        # Verifica se há pelo menos uma camada em um dos listWidgets
        if self.listWidgetEstacas.count() > 0 or self.listWidgetTalude.count() > 0:
            # Se houver itens em qualquer um dos widgets, habilita o pushButtonJuntar
            self.pushButtonJuntar.setEnabled(True)
        else:
            # Se ambos os widgets estiverem vazios, desabilita o pushButtonJuntar
            self.pushButtonJuntar.setEnabled(False)

    def on_pushButtonLogo_clicked(self):
        """Método chamado ao clicar no botão para selecionar o logo."""
        # Abrir um diálogo para selecionar a imagem
        logo_path, _ = QFileDialog.getOpenFileName(self, "Selecione o logo", "", "Image Files (*.png *.jpg *.jpeg)")

        if logo_path:
            # Armazenar o caminho da imagem selecionada na memória
            self.logo_path = logo_path
            self.mostrar_mensagem(f"Logotipo selecionado: {logo_path}", "Sucesso")

            # Atualizar o texto e a cor do botão, sem alterar o estilo original
            self.pushButtonLogo.setText("Logo ✓")
            self.pushButtonLogo.setStyleSheet(self.pushButtonLogo.styleSheet() + "color: blue;")
        else:
            self.mostrar_mensagem("Nenhum logotipo selecionado.", "Erro")

            # Resetar o botão para o estado inicial se não houver logo
            self.logo_path = None
            self.pushButtonLogo.setText("Logo")
            self.pushButtonLogo.setStyleSheet(self.pushButtonLogo.styleSheet() + "color: black;")

    def check_logo_status(self):
        """Verifica se existe um logo armazenado e ajusta o botão pushButtonLogo adequadamente."""
        if hasattr(self, 'logo_path') and self.logo_path:
            # Se houver um logo armazenado, exibe o texto "Logo ✓" e estilo azul
            self.pushButtonLogo.setText("Logo ✓")
            self.pushButtonLogo.setStyleSheet(self.pushButtonLogo.styleSheet() + "color: blue;")
        else:
            # Caso contrário, exibe o texto "Logo" com o estilo padrão
            self.pushButtonLogo.setText("Logo")
            self.pushButtonLogo.setStyleSheet(self.pushButtonLogo.styleSheet() + "color: black;")

    def on_pushButtonVertices_clicked(self):
        """
        Método para ser chamado ao clicar no botão "Vertices".
        Verifica se há widgets adicionados na scrollArea e ativa o botão "Calcular" se houver.
        """
        # Verifique se a scrollArea tem um layout configurado
        layout = self.scrollAreaWidgetContents.layout()

        if layout and layout.count() > 0:
            # Ativa o botão Calcular se houver widgets
            self.pushButtonCalcular.setEnabled(True)
        else:
            # Desativa o botão Calcular se não houver widgets
            self.pushButtonCalcular.setEnabled(False)

    def connect_selection_signal(self):
        """Conecta o sinal selectionChanged para detectar mudanças na seleção de feições e controlar o botão."""
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)

        if selected_layer and isinstance(selected_layer, QgsVectorLayer):
            selected_layer.selectionChanged.connect(self.on_selection_changed)

    def on_selection_changed(self):
        """Função chamada quando a seleção de feições é alterada. Atualiza o polígono e controla o estado do botão."""
        self.display_polygon()  # Atualiza a exibição do polígono
        selected_polygon_id = self.comboBoxPoligono.currentData()
        selected_layer = QgsProject.instance().mapLayer(selected_polygon_id)
        
        # Verifica se há feições selecionadas
        if selected_layer and selected_layer.selectedFeatureCount() > 0:
            self.pushButtonVertices.setEnabled(True)  # Ativa o botão
        else:
            self.pushButtonVertices.setEnabled(False)  # Desativa o botão

    def open_excel_with_attributes(self, item):
        """
        Abre um arquivo Excel com as informações da tabela de atributos da camada selecionada.
        """
        if not item or not item.text():
            self.mostrar_mensagem("Nenhuma camada selecionada.", "Erro")
            return

        # Obtém o nome da camada a partir do texto do item selecionado
        layer_name = item.text()
        layer = QgsProject.instance().mapLayersByName(layer_name)

        if not layer:
            self.mostrar_mensagem(f"Camada '{layer_name}' não encontrada.", "Erro")
            return

        layer = layer[0]

        if not isinstance(layer, QgsVectorLayer):
            self.mostrar_mensagem(f"'{layer_name}' não é uma camada de vetor.", "Erro")
            return

        # Extrai os atributos da camada
        fields = [field.name() for field in layer.fields()]
        features = [feat for feat in layer.getFeatures()]

        # Verifica se há feições na camada
        if not features:
            self.mostrar_mensagem(f"A camada '{layer_name}' não contém feições.", "Erro")
            return

        # Cria uma lista para armazenar os dados
        dados_tabela = []

        # Percorre todas as feições e campos para extrair os dados
        for feature in features:
            linha_dados = []
            for field_name in fields:
                valor_celula = feature[field_name]
                # Tenta converter para número se possível
                try:
                    valor_celula = float(valor_celula)
                except (ValueError, TypeError):
                    pass  # Mantém o valor como string se não for um número
                linha_dados.append(valor_celula)
            dados_tabela.append(linha_dados)

        # Cria um DataFrame a partir dos dados
        df = pd.DataFrame(dados_tabela, columns=fields)

        # Cria um arquivo temporário para o Excel
        temp_fd, fileName = tempfile.mkstemp(suffix=".xlsx")
        os.close(temp_fd)  # Fecha o arquivo temporário para evitar conflito de acesso

        try:
            # Salva o DataFrame no arquivo Excel
            df.to_excel(fileName, index=False)

            # Adiciona bordas a todas as células no Excel
            self.add_borders_to_excel(fileName)

            # Abre o Excel com o arquivo temporário
            os.startfile(fileName)  # Funciona no Windows
            self.mostrar_mensagem("A tabela foi exportada e aberta no Excel temporariamente.", "Sucesso")

        except Exception as e:
            self.mostrar_mensagem(f"Erro ao salvar o arquivo Excel: {str(e)}", "Erro")

    def add_borders_to_excel(self, file_path):
        """
        Adiciona bordas a todas as células do arquivo Excel especificado.
        """
        # Abre o workbook existente
        workbook = load_workbook(file_path)
        sheet = workbook.active

        # Define o estilo da borda
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Aplica a borda a todas as células usadas na planilha
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.border = thin_border

        # Ajusta a largura das colunas automaticamente
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Pega a letra da coluna
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width

        # Salva as alterações no arquivo
        workbook.save(file_path)
        workbook.close()

class GraficoManager:
    def __init__(self, plato_manager):
        """
        Inicializa o GraficoManager com uma referência para a classe PlatoManager.

        Parameters:
        - plato_manager: A instância de PlatoManager que contém o método mostrar_mensagem.
        """
        self.plato_manager = plato_manager
        self.logo_path = None  # Inicialmente, não há logo armazenado

    def find_layer(self, layer_name, layer_type):
        for layer in QgsProject.instance().mapLayers().values():
            if layer.name() == layer_name and isinstance(layer, layer_type):
                return layer
        return None

    def sample_raster_value(self, point, raster_layer):
        """Amostra o valor Z do raster baseado nas coordenadas do ponto."""
        identify_result = raster_layer.dataProvider().identify(QgsPointXY(point.x(), point.y()), QgsRaster.IdentifyFormatValue)
        
        if identify_result.isValid():
            results = identify_result.results()
            if results:
                band_key = list(results.keys())[0]
                z_value = results.get(band_key)  # Use get para evitar KeyError
                if z_value is not None:  # Verifica se o valor é válido
                    return round(float(z_value), 3)
        return None  # Retorna None se não houver valor válido

    def create_support_points_layer(self, estacas_layer, raster_layer):
        # Obter a resolução do pixel da camada raster
        raster_extent = raster_layer.extent()
        raster_width = raster_layer.width()
        raster_height = raster_layer.height()
        pixel_resolution_x = raster_extent.width() / raster_width
        pixel_resolution_y = raster_extent.height() / raster_height

        # Definir o espaçamento de suporte como a resolução do pixel
        support_spacing = min(pixel_resolution_x, pixel_resolution_y)

        # Obter o CRS da camada de estacas
        crs = estacas_layer.sourceCrs().authid()
        support_layer = QgsVectorLayer(f"Point?crs={crs}", "", "memory")  # Criar a camada sem nome

        prov = support_layer.dataProvider()

        # Adiciona os campos necessários, incluindo "Acumula_dist"
        prov.addAttributes([QgsField("ID", QVariant.Int), QgsField("Original_ID", QVariant.Int),
                            QgsField("X", QVariant.Double), QgsField("Y", QVariant.Double),
                            QgsField("Znovo", QVariant.Double), QgsField("Acumula_dist", QVariant.Double)])
        support_layer.updateFields()

        estacas_features = [feat for feat in estacas_layer.getFeatures()]
        estacas_points = [feat.geometry().asPoint() for feat in estacas_features]
        all_points = []
        support_point_id = 0
        last_coord = None
        acumula_dist = 0

        # Calcular o desnível para a extensão (1 metro além do desnível)
        first_desnivel = estacas_features[0]['Desnível']
        last_desnivel = estacas_features[-1]['Desnível']
        extend_by_start = min(abs(first_desnivel) + 3, 3)  # No máximo 10 metros
        extend_by_end = min(abs(last_desnivel) + 3, 3)  # No máximo 10 metros

        # Adicionar pontos extras antes do primeiro ponto de estacas
        first_segment_dir = estacas_points[1] - estacas_points[0]
        for i in range(int(extend_by_start // support_spacing), 0, -1):
            extra_point = QgsPointXY(estacas_points[0].x() - first_segment_dir.x() * (i * support_spacing) / first_segment_dir.length(),
                                     estacas_points[0].y() - first_segment_dir.y() * (i * support_spacing) / first_segment_dir.length())
            
            z_value = self.sample_raster_value(extra_point, raster_layer)
            if z_value is None:
                # Interrompe a extensão se o valor de Z for None
                break
            
            support_point_id += 1
            all_points.append((extra_point, -i, support_point_id))

        # Gerar os pontos de apoio ao longo dos segmentos
        for i, start_point in enumerate(estacas_points[:-1]):
            end_point = estacas_points[i + 1]
            segment_length = start_point.distance(end_point)
            num_intermediate_points = int(segment_length / support_spacing)

            for j in range(num_intermediate_points + 1):
                x = start_point.x() + (end_point.x() - start_point.x()) * (j * support_spacing) / segment_length
                y = start_point.y() + (end_point.y() - start_point.y()) * (j * support_spacing) / segment_length
                inter_point = QgsPointXY(x, y)
                support_point_id += 1
                all_points.append((inter_point, estacas_features[i]['ID'], support_point_id))

        # Adicionar pontos extras após o último ponto de estacas
        last_segment_dir = estacas_points[-1] - estacas_points[-2]
        for i in range(0, int(extend_by_end // support_spacing) + 1):
            extra_point = QgsPointXY(estacas_points[-1].x() + last_segment_dir.x() * (i * support_spacing) / last_segment_dir.length(),
                                     estacas_points[-1].y() + last_segment_dir.y() * (i * support_spacing) / last_segment_dir.length())
            
            z_value = self.sample_raster_value(extra_point, raster_layer)
            if z_value is None:
                # Interrompe a extensão se o valor de Z for None
                break

            support_point_id += 1
            all_points.append((extra_point, estacas_features[-1]['ID'], support_point_id))

        # Criar os recursos e adicionar à camada com a distância acumulada
        first_positive_id_encountered = False
        for point, original_id, support_point_id in all_points:
            current_coord = (point.x(), point.y())
            if not first_positive_id_encountered:
                if original_id >= 0:
                    acumula_dist = 0
                    first_positive_id_encountered = True
                else:
                    acumula_dist = -extend_by_start + (support_point_id - 1) * support_spacing
            else:
                if last_coord:
                    segment_distance = QgsPointXY(*last_coord).distance(QgsPointXY(*current_coord))
                    acumula_dist += segment_distance
            last_coord = current_coord

            # Atributos do recurso, incluindo acumulação de distância
            feat = QgsFeature()
            feat.setGeometry(QgsGeometry.fromPointXY(QgsPointXY(*current_coord)))
            feat.setAttributes([support_point_id, original_id, round(point.x(), 3), round(point.y(), 3), None, round(acumula_dist, 3)])
            prov.addFeature(feat)

         # Amostra os valores do MDT para os pontos de apoio e atualiza o campo "Znovo"
        support_layer.startEditing()
        z_value_previous = None

        for feature in support_layer.getFeatures():
            point = feature.geometry().asPoint()
            z_value = self.sample_raster_value(point, raster_layer)
            if z_value is None and z_value_previous is not None:
                z_value = z_value_previous
            if z_value is not None:
                z_value_previous = z_value
            feature['Znovo'] = round(z_value, 3) if z_value is not None else None
            support_layer.updateFeature(feature)

        support_layer.commitChanges()

        # Retorna a camada sem adicionar ao projeto
        return support_layer

    def find_slope_intersection(self, x, y, apoio_distances, apoio_elevations, angle_degrees):
        angle_radians = math.radians(angle_degrees)
        slope = math.tan(angle_radians)
        
        # Calcula o ponto final da linha de talude (apenas como referência se não houver interseção)
        x_end = x + math.cos(angle_radians)
        y_end = y + math.sin(angle_radians)

        # Encontrar a interseção com a linha "Znovo"
        for i in range(len(apoio_distances) - 1):
            x1, y1 = apoio_distances[i], apoio_elevations[i]
            x2, y2 = apoio_distances[i + 1], apoio_elevations[i + 1]

            if x1 == x2:  # Ignora segmentos verticais, pois não haverá interseção
                continue

            # Cálculo da inclinação e interceptação da linha entre os pontos de Znovo
            slope_znovo = (y2 - y1) / (x2 - x1)
            intercept_znovo = y1 - slope_znovo * x1

            # Calcula o ponto de interseção
            intersect_x = (intercept_znovo - y + slope * x) / (slope - slope_znovo)
            intersect_y = slope * (intersect_x - x) + y

            # Verifica se a interseção está dentro do segmento de Znovo
            if min(x1, x2) <= intersect_x <= max(x1, x2) and min(y1, y2) <= intersect_y <= max(y1, y2):
                return intersect_x, intersect_y

        # Se não houver interseção, retorna o ponto final presumido
        return x_end, y_end

    def plot_layers(self, estacas_layer, pontos_apoio_layer, angle_degrees_ti, angle_degrees_tf):
        """
        Exibe um gráfico com as camadas de estacas e pontos de apoio e desenha os taludes com base nos ângulos fornecidos.
        
        Parameters:
        - estacas_layer: Camada de estacas.
        - pontos_apoio_layer: Camada de pontos de apoio.
        - angle_degrees_ti: Ângulo do talude inicial.
        - angle_degrees_tf: Ângulo do talude final.
        """

        # Coletar dados das camadas
        estacas_data = [(f['Acumula_dist'], f['NovoZ'], f['Z'], f['Desnível'], f.geometry().asPoint().x(), f.geometry().asPoint().y()) 
                        for f in estacas_layer.getFeatures()]
        pontos_apoio_data = [(f['Acumula_dist'], f['Znovo']) for f in pontos_apoio_layer.getFeatures()]

        # Ordenar os dados
        estacas_data.sort(key=lambda x: x[0])
        pontos_apoio_data.sort(key=lambda x: x[0])

        # Desempacotar dados
        estacas_distances, estacas_novoz, z_values, desnivels, x_coords, y_coords = zip(*estacas_data)
        apoio_distances, apoio_elevations = zip(*pontos_apoio_data)

        # Criar o gráfico principal
        fig, ax = plt.subplots(figsize=(15, 9))

        # Ajusta o título da figura com o nome da camada
        nome_camada = estacas_layer.name()
        fig.canvas.manager.set_window_title(f'{nome_camada}')

        fig.subplots_adjust(left=0.08, right=0.98, bottom=0.12, top=0.9)

        ax.set_xlabel('Distância Acumulada (m)')
        ax.set_ylabel('Elevação (m)')
        ax.set_title('Perfil Longitudinal')

        # Plotar a linha "NovoZ"
        ax.plot(estacas_distances, estacas_novoz, color='blue', linewidth=2, label='Perfil do Platô')

        # Plotar a linha "Znovo"
        ax.plot(apoio_distances, apoio_elevations, color='orange', linewidth=1.5, label='Terreno')

        # Pegar os pontos iniciais e finais da linha "NovoZ"
        start_x, start_y = estacas_distances[0], estacas_novoz[0]
        end_x, end_y = estacas_distances[-1], estacas_novoz[-1]

        # Elevações correspondentes em Znovo
        start_znovo = apoio_elevations[0]
        end_znovo = apoio_elevations[-1]

        # Ajustar o ângulo de inclinação para o talude de entrada (TI)
        if start_y > start_znovo:
            angle_degrees_ti = angle_degrees_ti
        else:
            angle_degrees_ti = 180 - angle_degrees_ti

        # Ajustar o ângulo de inclinação para o talude de saída (TF)
        if end_y > end_znovo:
            angle_degrees_tf = 180 - angle_degrees_tf
        else:
            angle_degrees_tf = angle_degrees_tf

        # Calcular e plotar a linha de talude para o ponto inicial (TI)
        talude_start_intersection = self.find_slope_intersection(start_x, start_y, apoio_distances, apoio_elevations, angle_degrees_ti)
        if talude_start_intersection:
            if estacas_novoz[0] > start_znovo:
                ti_color = 'blue'
            else:
                ti_color = 'red'
            ax.plot([start_x, talude_start_intersection[0]], [start_y, talude_start_intersection[1]], linestyle='--', linewidth=2.5, color=ti_color)

        # Calcular e plotar a linha de talude para o ponto final (TF)
        talude_end_intersection = self.find_slope_intersection(end_x, end_y, apoio_distances, apoio_elevations, angle_degrees_tf)
        if talude_end_intersection:
            if estacas_novoz[-1] > end_znovo:
                tf_color = 'blue'
            else:
                tf_color = 'red'
            ax.plot([end_x, talude_end_intersection[0]], [end_y, talude_end_intersection[1]], linestyle='--', linewidth=2.5, color=tf_color)

        # Adicionar linhas verticais nas extremidades da linha Talude
        if talude_start_intersection:
            ax.vlines(x=talude_start_intersection[0], ymin=0, ymax=talude_start_intersection[1], colors='grey', linestyles='dashed', linewidth=0.5)

        if talude_end_intersection:
            ax.vlines(x=talude_end_intersection[0], ymin=0, ymax=talude_end_intersection[1], colors='grey', linestyles='dashed', linewidth=0.5)

        # Criar os valores combinados de X para interpolação
        corte_adjusted_x = [talude_start_intersection[0]] + list(estacas_distances) + [talude_end_intersection[0]]
        corte_adjusted_y = [talude_start_intersection[1]] + list(estacas_novoz) + [talude_end_intersection[1]]

        # Preparar as listas estendidas de x e y para 'Znovo' e 'Corte Ajustado'
        x_combined = sorted(set(corte_adjusted_x + list(apoio_distances)))
        y_corte_interpolated = np.interp(x_combined, corte_adjusted_x, corte_adjusted_y)
        y_znovo_interpolated = np.interp(x_combined, apoio_distances, apoio_elevations)

        # Limitar o intervalo de x para o preenchimento
        x_fill_min = talude_start_intersection[0]
        x_fill_max = talude_end_intersection[0]

        # Ajustes nos limites do eixo Y
        y_min0 = min(apoio_elevations) - 2  # Valor mínimo da elevação para a linha de base com uma margem
        y_min1 = min(min(estacas_novoz), y_min0) - 3  # Valor mínimo com margem adicional
        y_max = max(max(estacas_novoz), max(apoio_elevations)) + 3  # Valor máximo com margem adicional
        ax.set_ylim(y_min1, y_max)

        # Plotar a linha "Znovo"
        # ax.plot(apoio_distances, apoio_elevations, color='orange', linewidth=1.5, label='Znovo (Terreno Natural)')

        # Preencher abaixo da linha "Terreno Natural" até a linha de base 'y_min0'
        ax.fill_between(apoio_distances, apoio_elevations, y_min0, color='lightgreen', alpha=0.3, hatch='***')

        # Filtrar os valores x e y para o intervalo de interesse
        x_fill = [x for x in x_combined if x_fill_min <= x <= x_fill_max]
        y_corte_fill = np.interp(x_fill, corte_adjusted_x, corte_adjusted_y)
        y_znovo_fill = np.interp(x_fill, apoio_distances, apoio_elevations)

        # Preenchimento vermelho abaixo de "Znovo" e acima de "Corte Ajustado"
        ax.fill_between(x_fill, y_corte_fill, y_znovo_fill, where=(y_corte_fill < y_znovo_fill), 
                        interpolate=True, color='red', alpha=0.5, hatch='///', label='Área Cortada')

        # Preenchimento azul acima de "Znovo" e abaixo de "Corte Ajustado"
        ax.fill_between(x_fill, y_znovo_fill, y_corte_fill, where=(y_corte_fill > y_znovo_fill), 
                        interpolate=True, color='blue', alpha=0.5, hatch='\\\\', label='Área Aterrada')

        # Adicionar a legenda para Talude Cortado e Aterrado
        ax.plot([], [], linestyle='--', linewidth=2.5, color='red', label='Talude Cortado')
        ax.plot([], [], linestyle='--', linewidth=2.5, color='blue', label='Talude Aterrado')

        # Função auxiliar para adicionar textos ao gráfico
        def add_text(ax, text, x_pos, y_pos, text_color, bbox_props):
            """Adiciona texto ao gráfico."""
            ax.text(x_pos, y_pos, text, ha="left", va="bottom", transform=ax.transAxes,
                    fontsize=9, weight='bold', fontstyle='italic', color=text_color, bbox=bbox_props)

        # Propriedades da caixa de anotação
        bbox_props = dict(facecolor='white', alpha=1, boxstyle="round,pad=0.5")

        # Calcular as áreas cortada e aterrada (manualmente)
        area_cortada = np.trapz(y_znovo_fill[y_znovo_fill > y_corte_fill] - y_corte_fill[y_znovo_fill > y_corte_fill], x=np.array(x_fill)[y_znovo_fill > y_corte_fill])
        area_aterrada = np.trapz(y_corte_fill[y_corte_fill > y_znovo_fill] - y_znovo_fill[y_corte_fill > y_znovo_fill], x=np.array(x_fill)[y_corte_fill > y_znovo_fill])

        # Calcular as inclinações
        inclinacao_novoz = (estacas_novoz[-1] - estacas_novoz[0]) / (estacas_distances[-1] - estacas_distances[0])
        inclinacao_znovo = (apoio_elevations[-1] - apoio_elevations[0]) / (apoio_distances[-1] - apoio_distances[0])

        # Adicionar textos ao gráfico
        add_text(ax, f'Área de Aterro: {abs(area_aterrada):.2f} m²', 0.85, 1.06, 'blue', bbox_props)
        add_text(ax, f'Área de Corte: {abs(area_cortada):.2f} m²', 0.85, 1.016, 'magenta', bbox_props)
        add_text(ax, f'Inclinação do Corte/Aterro: {inclinacao_novoz * 100:.3f}%', 0.005, 1.016, 'blue', bbox_props)
        add_text(ax, f'Inclinação média do Terreno: {inclinacao_znovo * 100:.3f}%', 0.005, 1.06, 'orange', bbox_props)

        # Adicionar anotações e linhas verticais para 'Desnivel' sobre a linha "NovoZ"
        for dist, corte, z, desnivel, x, y in estacas_data:
            # Definir a cor com base no valor de 'Desnivel'
            if desnivel > 0:
                color = 'blue'
            elif desnivel < 0:
                color = 'red'
            else:
                color = 'black'
            
            # Adicionar a anotação com a cor correspondente e uma caixa ao redor do texto
            ax.annotate(f'{desnivel}', xy=(dist, corte), textcoords="offset points",
                        xytext=(0, 7), ha='center', va='bottom', color=color, weight='bold',
                        fontsize=8, bbox=dict(boxstyle="round,pad=0.3", facecolor='white', edgecolor="black", alpha=1))

            # Desenhar a linha vertical da linha "NovoZ" até o Eixo X
            ax.vlines(x=dist, ymin=0, ymax=corte, colors='grey', linestyles='dashed', linewidth=0.5)

        # Coletar todas as posições x onde as linhas verticais tocam o eixo X
        vertical_line_positions = list(estacas_distances)

        if talude_start_intersection:
            vertical_line_positions.append(talude_start_intersection[0])

        if talude_end_intersection:
            vertical_line_positions.append(talude_end_intersection[0])

        # Criar uma lista de xticks incluindo as posições das linhas verticais
        xticks = sorted(set(vertical_line_positions))

        # Formatar os rótulos dos xticks
        formatted_labels = [f"{tick:.0f}" for tick in xticks]

        # Configurar os xticks e seus rótulos
        ax.set_xticks(xticks)
        ax.set_xticklabels(formatted_labels, fontsize=8)

        # Personalizar os rótulos do eixo x
        for i, label in enumerate(ax.get_xticklabels()):
            if xticks[i] in vertical_line_positions:
                # Rótulos onde as linhas verticais tocam o eixo x em negrito e itálico
                label.set_fontweight('bold')
                label.set_fontstyle('italic')

        # Posicionar a legenda no canto superior esquerdo
        ax.legend(loc='upper left')
        ax.grid(axis='y', color='gray', alpha=0.5, linestyle='-', linewidth=0.4)

        # Configurar os ticks do eixo Y para variar a cada 1 metro
        ax.yaxis.set_major_locator(plt.MultipleLocator(1))

        # Definir uma altura fixa para as anotações em relação ao eixo y do gráfico
        altura_fixa_para_anotacoes = y_min1 + (y_max - y_min1) * 0.1 # 5% acima do limite inferior

        # Adicionar anotações para cada ponto na linha de "NovoZ"
        for dist, corte, z, desnivel in zip(estacas_distances, estacas_novoz, z_values, desnivels):
            # Determinar a cor da anotação com base no valor de "Desnível"
            if desnivel < 0:
                cor_anotacao = 'red'
            elif desnivel > 0:
                cor_anotacao = 'blue'
            else:
                cor_anotacao = 'black'

            # Anotação para "NovoZ" (Corte/Aterro)
            ax.annotate(f'{corte:.3f} m',
                        xy=(dist, altura_fixa_para_anotacoes),
                        xytext=(8, 0),
                        textcoords="offset points", ha='center', va='center', color=cor_anotacao, fontstyle='italic', fontweight='bold',
                        fontsize=8, rotation=90, bbox=dict(boxstyle="round,pad=0.3", facecolor='white', edgecolor=cor_anotacao))

            # Anotação para "Z" (Terreno Natural)
            ax.annotate(f'{z:.3f} m',
                        xy=(dist, altura_fixa_para_anotacoes),
                        xytext=(-10.5, 0),
                        textcoords="offset points", ha='left', va='center', color='black', fontstyle='italic', fontweight='bold',
                        fontsize=8, rotation=90, bbox=dict(boxstyle="round,pad=0.3", facecolor='white', edgecolor='orange'))

        # Adicionar rótulos representativos para "Z" e "NovoZ" alinhados com as anotações
        ax.text(0.042, (altura_fixa_para_anotacoes - y_min1) / (y_max - y_min1), 'Terreno Natural', ha='right', va='center',
                color='black', fontstyle='italic', fontweight='bold', fontsize=9, rotation=90,
                transform=ax.transAxes, bbox=dict(boxstyle="round,pad=0.3", facecolor='white', edgecolor='orange'))

        ax.text(0.06, (altura_fixa_para_anotacoes - y_min1) / (y_max - y_min1), 'Corte/Aterro', ha='right', va='center',
                color='blue', fontstyle='italic', fontweight='bold', fontsize=9, rotation=90,
                transform=ax.transAxes, bbox=dict(boxstyle="round,pad=0.5", facecolor='white', edgecolor='red'))

        # Exemplo de uso:
        fig = plt.gcf()
        self.adicionar_rosa_dos_ventos(fig, x_coords, y_coords)

        # Chama a função para adicionar o logotipo ao gráfico
        self.adicionar_logo(fig)

        plt.show()  # Exibe o gráfico

    def adicionar_rosa_dos_ventos(self, fig, x_coords, y_coords):
        """
        Adiciona uma rosa dos ventos ao gráfico.

        Parameters:
        - fig: a figura na qual a rosa dos ventos será desenhada.
        - x_coords: lista de coordenadas x do perfil.
        - y_coords: lista de coordenadas y do perfil.
        """
        # Defina o tamanho e a posição da Rosa dos Ventos
        radius = 0.035  # Raio do círculo (em unidades de figura)
        center = (0.50, 0.82)  # Posição do centro do círculo (em unidades de figura)

        # Desenha um círculo para a rosa dos ventos
        circle = Circle(center, radius, transform=fig.transFigure, edgecolor="black", facecolor="none", clip_on=False)
        fig.add_artist(circle)

        # Desenha linhas com setas para indicar as direções
        for angle in [0, 90, 180, 270]:
            x_start = center[0]
            y_start = center[1]
            x_end = center[0] + 1.0 * radius * np.cos(np.radians(angle))
            y_end = center[1] + 1.0 * radius * np.sin(np.radians(angle))

            # Adiciona uma seta ao final da linha
            arrowprops = dict(arrowstyle="->", lw=1.5, mutation_scale=10)
            fig.add_artist(patches.FancyArrowPatch((x_start, y_start), (x_end, y_end),
                                                   connectionstyle="arc3,rad=0.0",
                                                   transform=fig.transFigure, clip_on=False, **arrowprops))

        # Calcula o ângulo com base nas coordenadas fornecidas
        delta_x = x_coords[-1] - x_coords[0]
        delta_y = y_coords[-1] - y_coords[0]
        angle = math.atan2(delta_y, delta_x)

        # Converte o ângulo de radianos para graus para facilitar a interpretação
        angle_degrees = math.degrees(angle)

        # Determina a direção principal com base no ângulo
        if -45 < angle_degrees <= 45:
            directions = ["E", "N", "W", "S"]
        elif 45 < angle_degrees <= 135:
            directions = ["N", "W", "S", "E"]
        elif -135 < angle_degrees <= -45:
            directions = ["S", "E", "N", "W"]
        else:
            directions = ["W", "N", "E", "S"]

        # Adiciona rótulos para as direções
        label_distance = 1.4  # Ajuste essa variável para mudar a distância dos rótulos ao círculo
        for i, label in enumerate(directions):
            angle = 90 * i
            x = center[0] + label_distance * radius * np.cos(np.radians(angle))
            y = center[1] + label_distance * radius * np.sin(np.radians(angle))
            fig.text(x, y, label, ha="center", va="center")

    def adicionar_logo(self, fig):
        """
        Função para adicionar o logotipo ao gráfico, se houver um logotipo selecionado.
        
        A imagem é redimensionada proporcionalmente dentro de uma caixa de 250x150 pixels.
        
        Parameters:
        - fig: O objeto Figure do matplotlib onde o logotipo será adicionado.
        """
        if hasattr(self, 'logo_path') and self.logo_path:
            try:
                # Carregar o logotipo
                logo = plt.imread(self.logo_path)

                # Tamanho máximo da caixa onde o logo deve ser ajustado
                max_width, max_height = 150, 75

                # Obtenha a largura e a altura da imagem original
                img_height, img_width = logo.shape[:2]

                # Calcular o fator de escala para largura e altura
                width_scale = max_width / img_width
                height_scale = max_height / img_height

                # Escolher o menor fator de escala para garantir que a imagem caiba na caixa
                scale_factor = min(width_scale, height_scale)

                # Cria uma instância de OffsetImage com o logotipo redimensionado
                imagebox = OffsetImage(logo, zoom=scale_factor)

                # Opções personalizadas para a caixa
                bboxprops = dict(boxstyle="round,pad=0.5", facecolor="white", edgecolor="black", linewidth=1.5)

                # Cria uma instância de AnnotationBbox para posicionar o logotipo
                ab = AnnotationBbox(imagebox, (0.975, 0.89), frameon=True, xycoords='figure fraction',
                                    box_alignment=(1, 1), bboxprops=bboxprops)

                # Adiciona o logotipo ao gráfico
                ax = plt.gca()  # Obtém o eixo atual
                ax.add_artist(ab)

            except Exception as e:
                self.plato_manager.mostrar_mensagem(f"Erro ao carregar o logotipo: {str(e)}", "Erro")
        else:
            self.plato_manager.mostrar_mensagem("Nenhum logotipo disponível para adicionar ao gráfico.", "Sucesso")

class ListDeleteButtonDelegate(QStyledItemDelegate):
    def __init__(self, parent=None):
        super(ListDeleteButtonDelegate, self).__init__(parent)
        self.parent = parent

    def paint(self, painter, option, index):
        if index.isValid():
            # Ignora o cabeçalho (primeiro item)
            if index.row() == 0:
                super(ListDeleteButtonDelegate, self).paint(painter, option, index)
                return

            # Desenha o fundo com base no estado
            if option.state & QStyle.State_Selected:
                painter.fillRect(option.rect, option.palette.highlight())
            elif option.state & QStyle.State_MouseOver:
                # Estado de hover (passar o mouse sobre o item)
                hover_color = QColor("#00ffff")  # Cor definida no styleSheet
                painter.fillRect(option.rect, hover_color)
            else:
                painter.fillRect(option.rect, option.palette.base())

            # Calcula as posições
            rect = option.rect
            icon_size = 10
            icon_margin = 4
            icon_rect = QRect(
                rect.left() + icon_margin,
                rect.top() + (rect.height() - icon_size) // 2,
                icon_size,
                icon_size
            )
            text_rect = QRect(
                icon_rect.right() + icon_margin,
                rect.top(),
                rect.width() - icon_size - 3 * icon_margin,
                rect.height()
            )

            # Desenha o quadradinho com borda arredondada para exclusão
            painter.save()
            painter.setRenderHint(QPainter.Antialiasing)
            painter.setPen(QPen(QColor(0, 0, 255), 1))  # Cor da borda do quadrado
            painter.setBrush(QBrush(QColor(255, 0, 0, 200)))  # Fundo vermelho claro
            radius = 2  # Raio das bordas arredondadas
            painter.drawRoundedRect(icon_rect, radius, radius)  # Desenha o quadrado com bordas arredondadas

            # Desenha o "x" dentro do quadrado
            painter.setPen(QPen(QColor(255, 255, 255), 2))  # Cor e espessura do "x"
            painter.drawLine(icon_rect.topLeft() + QPoint(3, 3), icon_rect.bottomRight() - QPoint(3, 3))
            painter.drawLine(icon_rect.topRight() + QPoint(-3, 3), icon_rect.bottomLeft() + QPoint(3, -3))
            painter.restore()

            # Desenha o texto
            painter.save()
            # Ajusta a cor do texto com base no estado
            if option.state & QStyle.State_Selected:
                text_color = option.palette.highlightedText().color()
            else:
                text_color = option.palette.text().color()
            painter.setPen(text_color)
            font = painter.font()
            font.setPointSize(8)  # Ajusta o tamanho da fonte se necessário
            painter.setFont(font)
            text = index.data(Qt.DisplayRole)
            painter.drawText(text_rect, Qt.AlignVCenter | Qt.TextSingleLine, text)
            painter.restore()

    def editorEvent(self, event, model, option, index):
        if index.row() == 0:
            return False  # Ignora eventos no cabeçalho

        if event.type() == QEvent.MouseButtonRelease:
            # Calcula as posições
            rect = option.rect
            icon_size = 10
            icon_margin = 4
            icon_rect = QRect(
                rect.left() + icon_margin,
                rect.top() + (rect.height() - icon_size) // 2,
                icon_size,
                icon_size
            )
            if icon_rect.contains(event.pos()):
                # Remove a camada correspondente do QGIS e do listWidget
                layer_id = index.data(Qt.UserRole)  # Obtém o ID da camada armazenado no UserRole do item
                QgsProject.instance().removeMapLayer(layer_id)  # Remove a camada do QGIS
                # Não é necessário remover o item manualmente; o slot de atualização cuidará disso
                return True
        return super(ListDeleteButtonDelegate, self).editorEvent(event, model, option, index)
